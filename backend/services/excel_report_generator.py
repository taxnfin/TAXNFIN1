"""
excel_report_generator.py
Genera Excel corporativo con formato azul para Reporte Board.
backend/services/excel_report_generator.py
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta azul corporativo ──────────────────────────
NAVY    = "0D2B5E"
BLUE    = "1A56A8"
BLUE2   = "3B82F6"
BLUE_LT = "EFF6FF"
BLUE_MD = "DBEAFE"
WHITE   = "FFFFFF"
DGRAY   = "334155"
SLATE   = "64748B"
LGRAY   = "F8FAFC"
MGRAY   = "E2E8F0"
GREEN   = "059669"
GREEN_LT= "ECFDF5"
RED     = "DC2626"
RED_LT  = "FEF2F2"
AMBER   = "D97706"
AMBER_LT= "FFFBEB"

def _f(c): return PatternFill('solid', fgColor=c)
def _aln(h='left', v='center', w=False): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def _fnt(sz=9, bold=False, color=DGRAY): return Font(name='Arial', size=sz, bold=bold, color=color)
def _border(): 
    s = Side(style='thin', color=MGRAY)
    return Border(bottom=s)

def _sem(estado):
    e = str(estado or '').lower()
    if 'bueno' in e or 'good' in e: return GREEN, GREEN_LT
    if 'aten' in e or 'warn' in e:  return AMBER, AMBER_LT
    return RED, RED_LT

def _banner(ws, titulo, subtitulo, ncols=5):
    last = get_column_letter(ncols)
    ws.merge_cells(f'A1:{last}1')
    ws['A1'] = titulo
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color=WHITE)
    ws['A1'].fill = _f(NAVY)
    ws['A1'].alignment = _aln('center', 'center')
    ws.row_dimensions[1].height = 34

    ws.merge_cells(f'A2:{last}2')
    ws['A2'] = subtitulo
    ws['A2'].font = Font(name='Arial', size=9, color='93C5FD')
    ws['A2'].fill = _f(BLUE)
    ws['A2'].alignment = _aln('center', 'center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8

def _sec(ws, r, titulo, ncols=5):
    last = get_column_letter(ncols)
    ws.merge_cells(f'A{r}:{last}{r}')
    ws[f'A{r}'] = f'  {titulo}'
    ws[f'A{r}'].font = Font(name='Arial', size=10, bold=True, color=WHITE)
    ws[f'A{r}'].fill = _f(BLUE)
    ws[f'A{r}'].alignment = _aln('left', 'center')
    ws.row_dimensions[r].height = 22

def _col_hdr(ws, r, labels, height=16):
    for i, txt in enumerate(labels):
        c = ws.cell(row=r, column=i+1, value=txt)
        c.font = Font(name='Arial', size=8, bold=True, color=WHITE)
        c.fill = _f(NAVY)
        c.alignment = _aln('center')
    ws.row_dimensions[r].height = height

def build_excel_report(data: dict) -> io.BytesIO:
    empresa  = data.get('empresa', 'Empresa')
    periodo  = data.get('periodo', '')
    rfc      = data.get('rfc', '')
    inc      = data.get('income_statement', {})
    bal      = data.get('balance_sheet', {})
    met      = data.get('metrics', {})
    ai       = data.get('ai_analysis', {})
    trends   = data.get('trends', [])

    sub = f'{periodo}  ·  RFC: {rfc}  ·  TaxnFin · Claude Sonnet'

    wb = Workbook()

    # ══════════════════════════════════════════════════
    # HOJA 1 — RESUMEN EJECUTIVO
    # ══════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Resumen Ejecutivo'
    ws1.sheet_view.showGridLines = False
    for col, w in zip('ABCDE', [30, 20, 14, 4, 14]):
        ws1.column_dimensions[col].width = w

    _banner(ws1, f'REPORTE EJECUTIVO — {empresa.upper()}', sub, 5)

    # KPI cards
    ingresos    = inc.get('ingresos', 0) or 0
    util_bruta  = inc.get('utilidad_bruta', 0) or 0
    ebitda      = inc.get('ebitda', 0) or 0
    util_neta   = inc.get('utilidad_neta', 0) or 0

    kpis = [
        ('INGRESOS', ingresos, BLUE, BLUE_LT, '100%'),
        ('UTIL. BRUTA', util_bruta, GREEN, GREEN_LT, f'{(util_bruta/ingresos*100):.1f}%' if ingresos else ''),
        ('EBITDA', ebitda, RED if ebitda < 0 else GREEN, RED_LT if ebitda < 0 else GREEN_LT, f'{(ebitda/ingresos*100):.1f}%' if ingresos else ''),
        ('UTIL. NETA', util_neta, RED if util_neta < 0 else GREEN, RED_LT if util_neta < 0 else GREEN_LT, f'{(util_neta/ingresos*100):.1f}%' if ingresos else ''),
    ]
    for i, (lbl, v, vc, vbg, pct) in enumerate(kpis):
        col = get_column_letter(i + 1)
        ws1[f'{col}4'] = lbl
        ws1[f'{col}4'].font = _fnt(8, True, SLATE)
        ws1[f'{col}4'].fill = _f(LGRAY)
        ws1[f'{col}4'].alignment = _aln('center')
        ws1.row_dimensions[4].height = 15

        ws1[f'{col}5'] = v
        ws1[f'{col}5'].number_format = '$#,##0;($#,##0);"-"'
        ws1[f'{col}5'].font = Font(name='Arial', size=18, bold=True, color=vc)
        ws1[f'{col}5'].fill = _f(vbg)
        ws1[f'{col}5'].alignment = _aln('center')
        ws1.row_dimensions[5].height = 32

        ws1[f'{col}6'] = pct
        ws1[f'{col}6'].font = _fnt(9, False, SLATE)
        ws1[f'{col}6'].fill = _f(vbg)
        ws1[f'{col}6'].alignment = _aln('center')
        ws1.row_dimensions[6].height = 15

    ws1.row_dimensions[7].height = 8

    # Estado de Resultados
    r = 8
    _sec(ws1, r, 'ESTADO DE RESULTADOS', 5); r += 1
    _col_hdr(ws1, r, ['Concepto', 'Monto (MXN)', '% Ingresos', '', '']); r += 1

    gastos_op = (inc.get('gastos_venta', 0) or 0) + (inc.get('gastos_administracion', 0) or 0) + (inc.get('gastos_generales', 0) or 0)
    er_rows = [
        ('Ingresos',          ingresos,   f'{100:.1f}%',                             BLUE_LT, True),
        ('Costo de Ventas',   inc.get('costo_ventas', 0) or 0, f'{(inc.get("costo_ventas",0) or 0)/ingresos*100:.1f}%' if ingresos else '', None, False),
        ('Utilidad Bruta',    util_bruta, f'{util_bruta/ingresos*100:.1f}%' if ingresos else '', GREEN_LT, True),
        ('Gastos Operativos', gastos_op,  '',                                          None,    False),
        ('EBITDA',            ebitda,     f'{ebitda/ingresos*100:.1f}%' if ingresos else '', AMBER_LT if ebitda < 0 else GREEN_LT, True),
        ('Utilidad Neta',     util_neta,  f'{util_neta/ingresos*100:.1f}%' if ingresos else '', RED_LT if util_neta < 0 else GREEN_LT, True),
    ]
    for i, (concepto, v, pct, bg_fixed, bold) in enumerate(er_rows):
        bg = bg_fixed or (LGRAY if i % 2 == 0 else WHITE)
        v = v or 0
        color_v = RED if v < 0 else (GREEN if bold else DGRAY)
        ws1[f'A{r}'] = concepto
        ws1[f'A{r}'].font = _fnt(9, bold, NAVY if bold else DGRAY)
        ws1[f'A{r}'].fill = _f(bg); ws1[f'A{r}'].alignment = _aln('left', 'center')
        ws1[f'B{r}'] = v
        ws1[f'B{r}'].number_format = '$#,##0;($#,##0);"-"'
        ws1[f'B{r}'].font = _fnt(9, bold, color_v)
        ws1[f'B{r}'].fill = _f(bg); ws1[f'B{r}'].alignment = _aln('right', 'center')
        ws1[f'C{r}'] = pct
        ws1[f'C{r}'].font = _fnt(9, False, SLATE)
        ws1[f'C{r}'].fill = _f(bg); ws1[f'C{r}'].alignment = _aln('center', 'center')
        ws1.row_dimensions[r].height = 18; r += 1

    ws1.row_dimensions[r].height = 8; r += 1
    _sec(ws1, r, 'BALANCE GENERAL', 5); r += 1
    _col_hdr(ws1, r, ['Concepto', 'Monto (MXN)', '', '', '']); r += 1

    activo_total  = bal.get('activo_total', 0) or 0
    pasivo_total  = bal.get('pasivo_total', 0) or 0
    capital       = bal.get('capital_contable', 0) or 0
    bal_rows = [
        ('Activo Total',     activo_total,                       BLUE_LT, True),
        ('Activo Circulante',bal.get('activo_circulante', 0) or 0, None,  False),
        ('Activo Fijo',      bal.get('activo_fijo', 0) or 0,      None,  False),
        ('Pasivo Total',     pasivo_total,                       RED_LT,  True),
        ('Capital Contable', capital,                            GREEN_LT if capital >= 0 else RED_LT, True),
    ]
    for i, (concepto, v, bg_fixed, bold) in enumerate(bal_rows):
        bg = bg_fixed or (LGRAY if i % 2 == 0 else WHITE)
        v = v or 0
        ws1[f'A{r}'] = concepto
        ws1[f'A{r}'].font = _fnt(9, bold, NAVY if bold else DGRAY)
        ws1[f'A{r}'].fill = _f(bg); ws1[f'A{r}'].alignment = _aln('left', 'center')
        ws1[f'B{r}'] = v
        ws1[f'B{r}'].number_format = '$#,##0;($#,##0);"-"'
        ws1[f'B{r}'].font = _fnt(9, bold, RED if v < 0 else DGRAY)
        ws1[f'B{r}'].fill = _f(bg); ws1[f'B{r}'].alignment = _aln('right', 'center')
        ws1.row_dimensions[r].height = 18; r += 1

    # ══════════════════════════════════════════════════
    # HOJA 2 — DASHBOARD KPIs
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet('Dashboard KPIs')
    ws2.sheet_view.showGridLines = False
    for col, w in zip('ABCD', [28, 14, 14, 12]):
        ws2.column_dimensions[col].width = w
    _banner(ws2, f'DASHBOARD DE KPIs — {empresa.upper()}', sub, 4)

    margins  = met.get('margins', {})
    returns  = met.get('returns', {})
    liq      = met.get('liquidity', {})
    eff      = met.get('efficiency', {})
    solv     = met.get('solvency', {})

    def kpi_row(ws, r, nombre, valor_str, estado, i):
        bg = LGRAY if i % 2 == 0 else WHITE
        ec, ebg = _sem(estado)
        ws[f'A{r}'] = nombre
        ws[f'A{r}'].font = _fnt(9, False, DGRAY); ws[f'A{r}'].fill = _f(bg); ws[f'A{r}'].alignment = _aln('left', 'center')
        ws[f'B{r}'] = valor_str
        ws[f'B{r}'].font = Font(name='Arial', size=9, bold=True, color=ec); ws[f'B{r}'].fill = _f(bg); ws[f'B{r}'].alignment = _aln('center', 'center')
        ws[f'C{r}'] = estado or ''
        ws[f'C{r}'].font = Font(name='Arial', size=9, bold=True, color=ec); ws[f'C{r}'].fill = _f(ebg); ws[f'C{r}'].alignment = _aln('center', 'center')
        ws.row_dimensions[r].height = 18

    def get_v(d, key): return d.get(key, {}).get('value', 0) or 0
    def sem_m(v, g, w): return 'Bueno' if v >= g else ('Atención' if v >= w else 'Crítico')
    def sem_i(v, g, w): return 'Bueno' if v <= g else ('Atención' if v <= w else 'Crítico')

    r2 = 4
    secciones = [
        ('MÁRGENES', [
            ('Margen Bruto',    f'{get_v(margins,"gross_margin"):.1f}%',   sem_m(get_v(margins,"gross_margin"),30,15)),
            ('Margen EBITDA',   f'{get_v(margins,"ebitda_margin"):.1f}%',  sem_m(get_v(margins,"ebitda_margin"),10,0)),
            ('Margen Operativo',f'{get_v(margins,"operating_margin"):.1f}%',sem_m(get_v(margins,"operating_margin"),5,0)),
            ('Margen Neto',     f'{get_v(margins,"net_margin"):.1f}%',     sem_m(get_v(margins,"net_margin"),3,0)),
        ]),
        ('RETORNOS', [
            ('ROIC', f'{get_v(returns,"roic"):.1f}%', sem_m(get_v(returns,"roic"),8,3)),
            ('ROE',  f'{get_v(returns,"roe"):.1f}%',  sem_m(get_v(returns,"roe"),10,5)),
            ('ROCE', f'{get_v(returns,"roce"):.1f}%', sem_m(get_v(returns,"roce"),6,3)),
            ('ROA',  f'{get_v(returns,"roa"):.1f}%',  sem_m(get_v(returns,"roa"),5,2)),
        ]),
        ('LIQUIDEZ', [
            ('Razón Circulante', f'{get_v(liq,"current_ratio"):.2f}x',  sem_m(get_v(liq,"current_ratio"),1.5,1.0)),
            ('Prueba Ácida',     f'{get_v(liq,"quick_ratio"):.2f}x',    sem_m(get_v(liq,"quick_ratio"),1.0,0.5)),
            ('Razón Efectivo',   f'{get_v(liq,"cash_ratio"):.2f}x',     sem_m(get_v(liq,"cash_ratio"),0.2,0.1)),
            ('Cash Runway',      f'{get_v(liq,"cash_runway"):.1f}m',    sem_m(get_v(liq,"cash_runway"),3,1)),
        ]),
        ('CICLO OPERATIVO', [
            ('DSO Cobro',    f'{get_v(eff,"dso"):.0f}d',                  sem_i(get_v(eff,"dso"),60,90)),
            ('DIO Inventario',f'{get_v(eff,"dio"):.0f}d',                 sem_i(get_v(eff,"dio"),90,120)),
            ('DPO Pago',     f'{get_v(eff,"dpo"):.0f}d',                  sem_m(get_v(eff,"dpo"),45,30)),
            ('CCE',          f'{get_v(eff,"cash_conversion_cycle"):.0f}d',sem_i(get_v(eff,"cash_conversion_cycle"),90,120)),
        ]),
        ('SOLVENCIA', [
            ('Deuda/Capital', f'{get_v(solv,"debt_to_equity"):.2f}x',  sem_i(get_v(solv,"debt_to_equity"),1,2)),
            ('Deuda/Activos', f'{get_v(solv,"debt_to_assets"):.1f}%',  sem_i(get_v(solv,"debt_to_assets"),40,60)),
            ('Deuda/EBITDA',  f'{get_v(solv,"debt_to_ebitda"):.2f}x',  sem_i(get_v(solv,"debt_to_ebitda"),3,5)),
            ('Cobertura Int.',f'{get_v(solv,"interest_coverage"):.2f}x',sem_m(get_v(solv,"interest_coverage"),5,2)),
        ]),
    ]

    for sec_name, rows in secciones:
        _sec(ws2, r2, sec_name, 4); r2 += 1
        _col_hdr(ws2, r2, ['Indicador', 'Valor', 'Estado', '']); r2 += 1
        for i, (n, v, e) in enumerate(rows):
            kpi_row(ws2, r2, n, v, e, i); r2 += 1
        ws2.row_dimensions[r2].height = 6; r2 += 1

    # ══════════════════════════════════════════════════
    # HOJA 3 — ANÁLISIS IA
    # ══════════════════════════════════════════════════
    if ai:
        ws3 = wb.create_sheet('Análisis IA')
        ws3.sheet_view.showGridLines = False
        ws3.column_dimensions['A'].width = 28
        ws3.column_dimensions['B'].width = 80
        _banner(ws3, f'ANÁLISIS INTELIGENTE — {empresa.upper()}', sub, 2)
        _col_hdr(ws3, 4, ['Sección', 'Análisis']); r3 = 5
        ai_rows = [
            ('Resumen Ejecutivo',      ai.get('executive_summary', '')),
            ('Análisis de Rentabilidad',ai.get('profitability_analysis', '')),
            ('Análisis de Retornos',   ai.get('returns_analysis', '')),
            ('Análisis de Liquidez',   ai.get('liquidity_analysis', '')),
            ('Análisis de Solvencia',  ai.get('solvency_analysis', '')),
            ('Recomendaciones',        ai.get('recommendations', '')),
        ]
        for i, (sec, txt) in enumerate(ai_rows):
            bg = BLUE_LT if i % 2 == 0 else WHITE
            ws3[f'A{r3}'] = sec
            ws3[f'A{r3}'].font = _fnt(9, True, NAVY); ws3[f'A{r3}'].fill = _f(bg); ws3[f'A{r3}'].alignment = _aln('left', 'top')
            ws3[f'B{r3}'] = txt
            ws3[f'B{r3}'].font = _fnt(9, False, DGRAY); ws3[f'B{r3}'].fill = _f(bg); ws3[f'B{r3}'].alignment = _aln('left', 'top', True)
            ws3.row_dimensions[r3].height = max(40, min(120, len(str(txt)) // 2))
            r3 += 1

    # ══════════════════════════════════════════════════
    # HOJA 4 — TENDENCIAS
    # ══════════════════════════════════════════════════
    if trends:
        ws4 = wb.create_sheet('Tendencias')
        ws4.sheet_view.showGridLines = False
        for col, w in zip('ABCDEFGHI', [12, 16, 16, 14, 16, 12, 12, 12, 12]):
            ws4.column_dimensions[col].width = w
        _banner(ws4, f'TENDENCIAS HISTÓRICAS — {empresa.upper()}', sub, 9)
        hdrs = ['Período','Ingresos','Util. Bruta','EBITDA','Util. Neta','M. Bruto','M. Neto','ROE','ROIC']
        _col_hdr(ws4, 4, hdrs); r4 = 5
        for i, p in enumerate(trends):
            bg = BLUE_LT if i == 0 else (LGRAY if i % 2 == 0 else WHITE)
            inc_t = p.get('income_statement', {})
            met_t = p.get('metrics', {})
            row_data = [
                p.get('periodo',''),
                inc_t.get('ingresos', 0) or 0,
                inc_t.get('utilidad_bruta', 0) or 0,
                inc_t.get('ebitda', 0) or 0,
                inc_t.get('utilidad_neta', 0) or 0,
                f'{met_t.get("margins",{}).get("gross_margin",{}).get("value",0) or 0:.1f}%',
                f'{met_t.get("margins",{}).get("net_margin",{}).get("value",0) or 0:.1f}%',
                f'{met_t.get("returns",{}).get("roe",{}).get("value",0) or 0:.1f}%',
                f'{met_t.get("returns",{}).get("roic",{}).get("value",0) or 0:.1f}%',
            ]
            for j, v in enumerate(row_data):
                c = ws4.cell(row=r4, column=j+1, value=v)
                c.font = _fnt(9, j==0, NAVY if j==0 else DGRAY)
                c.fill = _f(bg); c.alignment = _aln('center', 'center')
                if j in [1,2,3,4]: c.number_format = '$#,##0;($#,##0);"-"'
            ws4.row_dimensions[r4].height = 18; r4 += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
