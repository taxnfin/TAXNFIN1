"""Cashflow weeks, transactions, manual projections and bank-transaction CRUD routes"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Dict, List, Optional, Any
from datetime import datetime, timezone, timedelta
import uuid
import io
import openpyxl
from openpyxl import Workbook

from core.database import db
from core.auth import get_current_user, get_active_company_id
from services.audit import audit_log
from services.fx import get_fx_rate_by_date
from services.cashflow import initialize_cashflow_weeks
from services.cfdi_parser import parse_cfdi_xml
from models.enums import UserRole, TransactionType, TransactionOrigin, BankTransactionType
from models.transaction import CashFlowWeek, Transaction, TransactionCreate
from models.bank import BankTransaction, BankTransactionCreate
from models.projection import ManualProjectionConcept, ManualProjectionConceptCreate
from models.cfdi import CFDI

router = APIRouter()

@router.get("/cashflow/weeks", response_model=List[CashFlowWeek])
async def get_cashflow_weeks(request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    weeks = await db.cashflow_weeks.find({'company_id': company_id}, {'_id': 0}).sort('fecha_inicio', 1).to_list(13)
    
    # If no weeks exist, generate them dynamically
    if not weeks:
        today = datetime.now(timezone.utc)
        start_of_current_week = today - timedelta(days=today.weekday())
        
        weeks = []
        for i in range(13):
            week_start = start_of_current_week + timedelta(weeks=i)
            week_end = week_start + timedelta(days=6, hours=23, minutes=59, seconds=59)
            weeks.append({
                'id': str(uuid.uuid4()),
                'company_id': company_id,
                'año': week_start.year,
                'numero_semana': week_start.isocalendar()[1],
                'fecha_inicio': week_start,
                'fecha_fin': week_end,
                'total_ingresos_reales': 0,
                'total_egresos_reales': 0,
                'total_ingresos_proyectados': 0,
                'total_egresos_proyectados': 0,
                'saldo_inicial': 0,
                'saldo_final_real': 0,
                'saldo_final_proyectado': 0,
                'created_at': today
            })
    
    # Get FX rates for conversion
    fx_rates = await db.fx_rates.find(
        {'company_id': company_id},
        {'_id': 0, 'moneda_origen': 1, 'moneda_destino': 1, 'tasa': 1}
    ).sort('fecha_vigencia', -1).to_list(100)
    
    # Build FX rates map
    fx_map = {'MXN': 1.0}
    for rate in fx_rates:
        if rate.get('moneda_destino') == 'MXN':
            fx_map[rate['moneda_origen']] = rate['tasa']
        elif rate.get('moneda_origen') == 'MXN':
            fx_map[rate['moneda_destino']] = 1 / rate['tasa']
    
    if 'USD' not in fx_map:
        fx_map['USD'] = 17.50
    if 'EUR' not in fx_map:
        fx_map['EUR'] = 19.00
    
    # Get initial balance from bank accounts with historical rates
    bank_accounts = await db.bank_accounts.find({'company_id': company_id, 'activo': True}, {'_id': 0}).to_list(100)
    saldo_inicial_total = 0.0
    for acc in bank_accounts:
        saldo = acc.get('saldo_inicial', 0)
        moneda = acc.get('moneda', 'MXN')
        fecha_saldo = acc.get('fecha_saldo')
        
        # Use historical rate if fecha_saldo is available
        if fecha_saldo:
            if isinstance(fecha_saldo, str):
                fecha_saldo = datetime.fromisoformat(fecha_saldo.replace('Z', '+00:00'))
            tasa = await get_fx_rate_by_date(company_id, moneda, fecha_saldo)
        else:
            tasa = fx_map.get(moneda, 1.0)
        
        saldo_inicial_total += saldo * tasa
    
    # Get CFDIs for calculating real inflows/outflows per week
    cfdis = await db.cfdis.find({'company_id': company_id}, {'_id': 0}).to_list(1000)
    
    # Track running balance
    running_balance = saldo_inicial_total
    
    for i, week in enumerate(weeks):
        for field in ['fecha_inicio', 'fecha_fin', 'created_at']:
            if isinstance(week.get(field), str):
                week[field] = datetime.fromisoformat(week[field].replace('Z', '+00:00'))
            # Ensure timezone aware
            if week.get(field) and week[field].tzinfo is None:
                week[field] = week[field].replace(tzinfo=timezone.utc)
        
        week_start = week['fecha_inicio']
        week_end = week['fecha_fin']
        
        # Calculate from CFDIs
        week_ingresos = 0
        week_egresos = 0
        for cfdi in cfdis:
            cfdi_date = cfdi.get('fecha_emision')
            if isinstance(cfdi_date, str):
                cfdi_date = datetime.fromisoformat(cfdi_date.replace('Z', '+00:00'))
            if cfdi_date and cfdi_date.tzinfo is None:
                cfdi_date = cfdi_date.replace(tzinfo=timezone.utc)
            
            if cfdi_date and week_start <= cfdi_date <= week_end:
                if cfdi.get('tipo_cfdi') == 'ingreso':
                    week_ingresos += cfdi.get('total', 0)
                else:
                    week_egresos += cfdi.get('total', 0)
        
        week['total_ingresos_reales'] = week_ingresos
        week['total_egresos_reales'] = week_egresos
        week['total_ingresos_proyectados'] = 0
        week['total_egresos_proyectados'] = 0
        
        # Set saldo_inicial from running balance
        week['saldo_inicial'] = running_balance
        
        # Calculate saldo_final
        week['saldo_final_real'] = week['saldo_inicial'] + week_ingresos - week_egresos
        week['saldo_final_proyectado'] = week['saldo_final_real']
        
        # Update running balance for next week
        running_balance = week['saldo_final_real']
    
    return weeks

@router.post("/transactions", response_model=Transaction)
async def create_transaction(transaction_data: TransactionCreate, request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    
    account = await db.bank_accounts.find_one({'id': transaction_data.bank_account_id, 'company_id': company_id}, {'_id': 0})
    if not account:
        raise HTTPException(status_code=404, detail="Cuenta bancaria no encontrada")
    
    transaction_date = transaction_data.fecha_transaccion
    week = await db.cashflow_weeks.find_one({
        'company_id': company_id,
        'fecha_inicio': {'$lte': transaction_date.isoformat()},
        'fecha_fin': {'$gte': transaction_date.isoformat()}
    }, {'_id': 0})
    
    if not week:
        raise HTTPException(status_code=400, detail="No se encontró semana de cashflow para la fecha")
    
    transaction = Transaction(
        company_id=company_id,
        cashflow_week_id=week['id'],
        **transaction_data.model_dump()
    )
    
    doc = transaction.model_dump()
    for field in ['fecha_transaccion', 'created_at']:
        doc[field] = doc[field].isoformat()
    await db.transactions.insert_one(doc)
    
    await audit_log(transaction.company_id, 'Transaction', transaction.id, 'CREATE', current_user['id'])
    return transaction

@router.get("/transactions", response_model=List[Transaction])
async def list_transactions(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0)
):
    company_id = await get_active_company_id(request, current_user)
    
    transactions = await db.transactions.find(
        {'company_id': company_id},
        {'_id': 0}
    ).sort('fecha_transaccion', -1).skip(skip).limit(limit).to_list(limit)
    
    for t in transactions:
        for field in ['fecha_transaccion', 'created_at']:
            if isinstance(t.get(field), str):
                t[field] = datetime.fromisoformat(t[field])
    return transactions

@router.post("/transactions/import")
async def import_transactions(file: UploadFile = File(...), current_user: Dict = Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel")
    
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    
    imported = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        
        try:
            transaction_data = TransactionCreate(
                bank_account_id=str(row[0]),
                concepto=str(row[1]),
                monto=float(row[2]),
                tipo_transaccion=TransactionType(row[3].lower()),
                fecha_transaccion=row[4] if isinstance(row[4], datetime) else datetime.fromisoformat(str(row[4])),
                es_real=bool(row[5]) if len(row) > 5 else False,
                es_proyeccion=bool(row[6]) if len(row) > 6 else True,
                vendor_id=str(row[7]) if len(row) > 7 and row[7] else None,
                customer_id=str(row[8]) if len(row) > 8 and row[8] else None
            )
            await create_transaction(transaction_data, current_user)
            imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {str(e)}")
    
    return {
        'status': 'success',
        'imported': imported,
        'errors': errors
    }

@router.post("/cfdi/upload")
async def upload_cfdi(request: Request, file: UploadFile = File(...), current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    
    if not file.filename.endswith('.xml'):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos XML")
    
    xml_content = await file.read()
    xml_str = xml_content.decode('utf-8')
    
    parsed = parse_cfdi_xml(xml_str)
    
    # Check for duplicate by exact UUID
    existing = await db.cfdis.find_one({'company_id': company_id, 'uuid': parsed['uuid']}, {'_id': 0})
    if existing:
        raise HTTPException(status_code=400, detail="CFDI ya existe (UUID duplicado)")
    
    # Also check for potential duplicate from Alegra by matching key fields
    # This catches cases where the same invoice was synced from Alegra with a pseudo-UUID
    fecha_emision_str = parsed['fecha_emision'][:10] if parsed['fecha_emision'] else ''
    potential_duplicate = await db.cfdis.find_one({
        'company_id': company_id,
        'emisor_rfc': parsed['emisor_rfc'],
        'receptor_rfc': parsed['receptor_rfc'],
        'total': parsed['total'],
        'fecha_emision': {'$regex': f'^{fecha_emision_str}'},
        'source': 'alegra'  # Only check Alegra records to avoid false positives
    }, {'_id': 0, 'id': 1, 'uuid': 1, 'referencia': 1, 'alegra_id': 1})
    
    if potential_duplicate:
        # Update the Alegra record with the real UUID from the XML
        await db.cfdis.update_one(
            {'id': potential_duplicate['id']},
            {'$set': {
                'uuid': parsed['uuid'],
                'xml_original': xml_str,
                'source': 'alegra+xml',  # Mark as merged
                'metodo_pago': parsed.get('metodo_pago', ''),
                'forma_pago': parsed.get('forma_pago', ''),
                'uso_cfdi': parsed.get('uso_cfdi', ''),
                'iva_trasladado': parsed.get('iva_trasladado', 0),
                'isr_retenido': parsed.get('isr_retenido', 0),
                'iva_retenido': parsed.get('iva_retenido', 0),
                'updated_at': datetime.now(timezone.utc).isoformat()
            }}
        )
        logger.info(f"CFDI de Alegra actualizado con UUID real: {parsed['uuid']} (Alegra ref: {potential_duplicate.get('referencia')})")
        
        # Return the updated CFDI
        updated_cfdi = await db.cfdis.find_one({'id': potential_duplicate['id']}, {'_id': 0, 'xml_original': 0})
        return {
            "success": True,
            "message": f"CFDI actualizado - ya existía de Alegra (Ref: {potential_duplicate.get('referencia', 'N/A')})",
            "cfdi": updated_cfdi,
            "merged_with_alegra": True
        }
    
    # Get company RFC to determine if this is income or expense
    company = await db.companies.find_one({'id': company_id}, {'_id': 0, 'rfc': 1, 'nombre': 1})
    company_rfc = company.get('rfc', '').upper().strip() if company else ''
    company_nombre = company.get('nombre', '').upper().strip() if company else ''
    emisor_rfc = parsed['emisor_rfc'].upper().strip()
    emisor_nombre = parsed['emisor_nombre'].upper().strip()
    
    # Clasificación de CFDI:
    # 1. Si es nómina/sueldos -> SIEMPRE es egreso (pago a empleados)
    # 2. Si el EMISOR es la empresa (mismo RFC o nombre) = INGRESO (la empresa emitió, cobrará)
    # 3. Si el EMISOR es diferente = EGRESO/GASTO (otra empresa emitió, la empresa pagará)
    
    if parsed.get('is_nomina'):
        # Nómina/sueldos SIEMPRE es egreso, incluso si el RFC es de la empresa
        tipo_cfdi = 'egreso'
        logger.info(f"CFDI {parsed['uuid']} clasificado como EGRESO (nómina/sueldos)")
    elif company_rfc and emisor_rfc == company_rfc:
        tipo_cfdi = 'ingreso'
    elif company_nombre and emisor_nombre == company_nombre:
        tipo_cfdi = 'ingreso'
    else:
        tipo_cfdi = 'egreso'
    
    cfdi = CFDI(
        company_id=company_id,
        uuid=parsed['uuid'],
        tipo_cfdi=CFDIType(tipo_cfdi),
        emisor_rfc=parsed['emisor_rfc'],
        emisor_nombre=parsed['emisor_nombre'],
        receptor_rfc=parsed['receptor_rfc'],
        receptor_nombre=parsed['receptor_nombre'],
        fecha_emision=datetime.fromisoformat(parsed['fecha_emision']),
        fecha_timbrado=datetime.fromisoformat(parsed['fecha_timbrado']),
        moneda=parsed['moneda'],
        subtotal=parsed['subtotal'],
        descuento=parsed.get('descuento', 0),
        impuestos=parsed['impuestos'],
        total=parsed['total'],
        metodo_pago=parsed.get('metodo_pago', ''),
        forma_pago=parsed.get('forma_pago', ''),
        uso_cfdi=parsed.get('uso_cfdi', ''),
        iva_trasladado=parsed.get('iva_trasladado', 0),
        isr_retenido=parsed.get('isr_retenido', 0),
        iva_retenido=parsed.get('iva_retenido', 0),
        ieps=parsed.get('ieps', 0),
        estado_cancelacion='vigente',
        xml_original=xml_str,
        source='xml'  # Mark as uploaded from XML
    )
    
    doc = cfdi.model_dump()
    for field in ['fecha_emision', 'fecha_timbrado', 'created_at']:
        doc[field] = doc[field].isoformat()
    await db.cfdis.insert_one(doc)
    
    await audit_log(cfdi.company_id, 'CFDI', cfdi.id, 'UPLOAD', current_user['id'])
    
    # Auto-categorize with AI if categories exist
    ai_category = None
    try:
        from ai_categorization_service import categorize_cfdi_with_ai
        
        # Get available categories for this CFDI type
        categories = await db.categories.find({
            'company_id': company_id, 
            'activo': True,
            'tipo': tipo_cfdi
        }, {'_id': 0}).to_list(100)
        
        if categories:
            # Get subcategories
            for cat in categories:
                subcats = await db.subcategories.find({'category_id': cat['id'], 'activo': True}, {'_id': 0}).to_list(100)
                cat['subcategorias'] = subcats
            
            # Prepare CFDI data for AI
            cfdi_data = {
                'uuid': cfdi.uuid,
                'tipo_cfdi': tipo_cfdi,
                'emisor_rfc': parsed['emisor_rfc'],
                'emisor_nombre': parsed['emisor_nombre'],
                'receptor_rfc': parsed['receptor_rfc'],
                'receptor_nombre': parsed['receptor_nombre'],
                'total': parsed['total'],
                'moneda': parsed['moneda'],
                'fecha_emision': parsed['fecha_emision']
            }
            
            # Call AI service
            ai_result = await categorize_cfdi_with_ai(cfdi_data, categories)
            
            if ai_result.get('success') and ai_result.get('category_id') and ai_result.get('confidence', 0) >= 70:
                update_data = {'category_id': ai_result['category_id']}
                if ai_result.get('subcategory_id'):
                    update_data['subcategory_id'] = ai_result['subcategory_id']
                
                await db.cfdis.update_one({'id': cfdi.id}, {'$set': update_data})
                ai_category = {
                    'category_id': ai_result['category_id'],
                    'subcategory_id': ai_result.get('subcategory_id'),
                    'confidence': ai_result.get('confidence'),
                    'reasoning': ai_result.get('reasoning')
                }
                await audit_log(company_id, 'CFDI', cfdi.id, 'AI_AUTO_CATEGORIZE', current_user['id'])
    except Exception as e:
        logger.warning(f"Auto-categorization failed for CFDI {cfdi.uuid}: {str(e)}")
    
    # NÓMINA: Auto-categorize as "Sueldos" and auto-reconcile with bank transactions
    nomina_auto_reconciled = None
    if parsed.get('is_nomina') or parsed.get('es_nomina_tipo_comprobante'):
        try:
            # 1. Find or create "Sueldos" category
            sueldos_category = await db.categories.find_one({
                'company_id': company_id,
                'nombre': {'$regex': '^sueldos?$', '$options': 'i'}
            }, {'_id': 0})
            
            if not sueldos_category:
                # Also check for "Nómina" or "Nominas"
                sueldos_category = await db.categories.find_one({
                    'company_id': company_id,
                    'nombre': {'$regex': '^n[oó]minas?$', '$options': 'i'}
                }, {'_id': 0})
            
            if not sueldos_category:
                # Create "Sueldos" category if doesn't exist
                new_cat_id = str(uuid.uuid4())
                sueldos_category = {
                    'id': new_cat_id,
                    'company_id': company_id,
                    'nombre': 'Sueldos',
                    'tipo': 'egreso',
                    'activo': True,
                    'created_at': datetime.now(timezone.utc).isoformat()
                }
                await db.categories.insert_one(sueldos_category)
                logger.info(f"Created 'Sueldos' category for company {company_id}")
            
            # Assign category to CFDI
            await db.cfdis.update_one({'id': cfdi.id}, {'$set': {'category_id': sueldos_category['id']}})
            
            # 2. Auto-reconcile with bank transactions by employee name and date
            receptor_nombre = parsed.get('receptor_nombre', '').upper().strip()
            fecha_emision = parsed.get('fecha_emision', '')[:10]  # YYYY-MM-DD
            nomina_data = parsed.get('nomina_data', {})
            fecha_pago_nomina = nomina_data.get('fecha_pago', fecha_emision)[:10] if nomina_data else fecha_emision
            total_cfdi = parsed.get('total', 0)
            
            # Search for bank transactions matching:
            # - Similar amount (within 5%)
            # - Similar date (within 7 days of fecha_pago)
            # - Name match in descripcion
            if receptor_nombre and total_cfdi > 0:
                # Parse dates for search range
                try:
                    fecha_ref = datetime.strptime(fecha_pago_nomina, '%Y-%m-%d')
                    fecha_desde = (fecha_ref - timedelta(days=7)).isoformat()
                    fecha_hasta = (fecha_ref + timedelta(days=7)).isoformat()
                    
                    # Find matching bank transactions
                    bank_txns = await db.bank_transactions.find({
                        'company_id': company_id,
                        'tipo_movimiento': 'debito',  # Payroll is a debit (money going out)
                        'conciliado': {'$ne': True},
                        'fecha_movimiento': {'$gte': fecha_desde, '$lte': fecha_hasta}
                    }, {'_id': 0}).to_list(100)
                    
                    best_match = None
                    best_score = 0
                    
                    for txn in bank_txns:
                        score = 0
                        txn_monto = abs(txn.get('monto', 0))
                        txn_desc = txn.get('descripcion', '').upper()
                        
                        # Check amount match (within 5%)
                        if txn_monto > 0:
                            diff_pct = abs(txn_monto - total_cfdi) / total_cfdi * 100
                            if diff_pct <= 5:
                                score += 50
                            elif diff_pct <= 10:
                                score += 30
                        
                        # Check name match in description
                        receptor_parts = receptor_nombre.split()
                        matches = sum(1 for part in receptor_parts if len(part) > 2 and part in txn_desc)
                        if matches >= 2:
                            score += 40
                        elif matches >= 1:
                            score += 20
                        
                        if score > best_score and score >= 50:
                            best_score = score
                            best_match = txn
                    
                    # Auto-reconcile if good match found
                    if best_match and best_score >= 50:
                        recon_id = str(uuid.uuid4())
                        recon_doc = {
                            'id': recon_id,
                            'company_id': company_id,
                            'bank_transaction_id': best_match['id'],
                            'cfdi_id': cfdi.id,
                            'metodo_conciliacion': 'auto_nomina',
                            'porcentaje_match': best_score,
                            'notas': f'Auto-conciliado: Nómina de {receptor_nombre}',
                            'fecha_conciliacion': datetime.now(timezone.utc).isoformat(),
                            'created_at': datetime.now(timezone.utc).isoformat()
                        }
                        await db.reconciliations.insert_one(recon_doc)
                        
                        # Update bank transaction as reconciled
                        await db.bank_transactions.update_one(
                            {'id': best_match['id']},
                            {'$set': {'conciliado': True, 'fecha_conciliacion': datetime.now(timezone.utc).isoformat()}}
                        )
                        
                        # Update CFDI as reconciled
                        await db.cfdis.update_one(
                            {'id': cfdi.id},
                            {'$set': {'estado_conciliacion': 'conciliado'}}
                        )
                        
                        nomina_auto_reconciled = {
                            'bank_transaction_id': best_match['id'],
                            'bank_descripcion': best_match.get('descripcion', '')[:50],
                            'match_score': best_score,
                            'empleado': receptor_nombre
                        }
                        
                        logger.info(f"Auto-reconciled nómina CFDI {cfdi.uuid} for {receptor_nombre} with bank txn {best_match['id']}")
                except Exception as inner_e:
                    logger.warning(f"Nómina auto-reconciliation failed: {str(inner_e)}")
        except Exception as e:
            logger.warning(f"Nómina processing failed for CFDI {cfdi.uuid}: {str(e)}")
    
    # Auto-detect customer/vendor by RFC
    auto_linked = None
    try:
        if tipo_cfdi == 'ingreso':
            # For income, the receptor (customer) is who we billed
            customer = await db.customers.find_one({
                'company_id': company_id, 
                'rfc': {'$regex': f'^{parsed["receptor_rfc"]}$', '$options': 'i'}
            }, {'_id': 0, 'id': 1, 'nombre': 1})
            if customer:
                await db.cfdis.update_one({'id': cfdi.id}, {'$set': {'customer_id': customer['id']}})
                auto_linked = {'type': 'customer', 'id': customer['id'], 'nombre': customer['nombre']}
        else:
            # For expense, the emisor (vendor) is who we're paying
            vendor = await db.vendors.find_one({
                'company_id': company_id, 
                'rfc': {'$regex': f'^{parsed["emisor_rfc"]}$', '$options': 'i'}
            }, {'_id': 0, 'id': 1, 'nombre': 1})
            if vendor:
                await db.cfdis.update_one({'id': cfdi.id}, {'$set': {'vendor_id': vendor['id']}})
                auto_linked = {'type': 'vendor', 'id': vendor['id'], 'nombre': vendor['nombre']}
    except Exception as e:
        logger.warning(f"Auto-link failed for CFDI {cfdi.uuid}: {str(e)}")
    
    # Detect new RFC - suggest creating customer/vendor
    new_rfc_detected = None
    if not auto_linked:
        try:
            if tipo_cfdi == 'ingreso':
                # Check if receptor RFC exists
                rfc_to_check = parsed.get('receptor_rfc', '')
                nombre_sugerido = parsed.get('receptor_nombre', rfc_to_check)
                existing = await db.customers.find_one({
                    'company_id': company_id,
                    'rfc': {'$regex': f'^{rfc_to_check}$', '$options': 'i'}
                }, {'_id': 0})
                if not existing and rfc_to_check:
                    new_rfc_detected = {
                        'type': 'customer',
                        'rfc': rfc_to_check,
                        'nombre_sugerido': nombre_sugerido,
                        'message': f'RFC {rfc_to_check} no está registrado como cliente. ¿Desea crearlo?'
                    }
            else:
                # Check if emisor RFC exists  
                rfc_to_check = parsed.get('emisor_rfc', '')
                nombre_sugerido = parsed.get('emisor_nombre', rfc_to_check)
                existing = await db.vendors.find_one({
                    'company_id': company_id,
                    'rfc': {'$regex': f'^{rfc_to_check}$', '$options': 'i'}
                }, {'_id': 0})
                if not existing and rfc_to_check:
                    new_rfc_detected = {
                        'type': 'vendor',
                        'rfc': rfc_to_check,
                        'nombre_sugerido': nombre_sugerido,
                        'message': f'RFC {rfc_to_check} no está registrado como proveedor. ¿Desea crearlo?'
                    }
        except Exception as e:
            logger.warning(f"New RFC detection failed: {str(e)}")
    
    return {
        'status': 'success', 
        'cfdi_id': cfdi.id, 
        'uuid': cfdi.uuid,
        'ai_categorized': ai_category is not None,
        'ai_category': ai_category,
        'auto_linked': auto_linked,
        'new_rfc_detected': new_rfc_detected,
        'is_nomina': parsed.get('is_nomina', False),
        'nomina_auto_reconciled': nomina_auto_reconciled
    }


@router.post("/cfdi/{cfdi_id}/link-xml")
async def link_xml_to_cfdi(
    cfdi_id: str,
    request: Request,
    file: UploadFile = File(...),
    current_user: Dict = Depends(get_current_user)
):
    """
    Link an XML file to an existing CFDI (typically from Alegra).
    This updates the CFDI with the real UUID from the XML and stores the XML content.
    """
    company_id = await get_active_company_id(request, current_user)
    
    # Find the existing CFDI
    existing_cfdi = await db.cfdis.find_one({'id': cfdi_id, 'company_id': company_id}, {'_id': 0})
    if not existing_cfdi:
        raise HTTPException(status_code=404, detail="CFDI no encontrado")
    
    if not file.filename.endswith('.xml'):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos XML")
    
    xml_content = await file.read()
    xml_str = xml_content.decode('utf-8')
    
    parsed = parse_cfdi_xml(xml_str)
    
    # Check if this UUID already exists in another CFDI
    uuid_exists = await db.cfdis.find_one({
        'company_id': company_id, 
        'uuid': parsed['uuid'],
        'id': {'$ne': cfdi_id}  # Exclude current CFDI
    }, {'_id': 0, 'id': 1})
    
    if uuid_exists:
        raise HTTPException(status_code=400, detail=f"El UUID {parsed['uuid']} ya existe en otro CFDI")
    
    # Update the CFDI with XML data
    old_source = existing_cfdi.get('source', 'unknown')
    new_source = f"{old_source}+xml" if old_source and old_source != 'unknown' else 'xml'
    
    update_data = {
        'uuid': parsed['uuid'],
        'xml_original': xml_str,
        'source': new_source,
        'metodo_pago': parsed.get('metodo_pago') or existing_cfdi.get('metodo_pago'),
        'forma_pago': parsed.get('forma_pago') or existing_cfdi.get('forma_pago'),
        'uso_cfdi': parsed.get('uso_cfdi') or existing_cfdi.get('uso_cfdi'),
        'iva_trasladado': parsed.get('iva_trasladado', 0),
        'isr_retenido': parsed.get('isr_retenido', 0),
        'iva_retenido': parsed.get('iva_retenido', 0),
        'updated_at': datetime.now(timezone.utc).isoformat()
    }
    
    await db.cfdis.update_one({'id': cfdi_id}, {'$set': update_data})
    
    logger.info(f"XML vinculado a CFDI {cfdi_id}: {existing_cfdi.get('uuid', 'N/A')} -> {parsed['uuid']}")
    
    # Get updated CFDI
    updated_cfdi = await db.cfdis.find_one({'id': cfdi_id}, {'_id': 0, 'xml_original': 0})
    
    return {
        "success": True,
        "message": f"XML vinculado exitosamente. UUID actualizado: {parsed['uuid']}",
        "cfdi": updated_cfdi,
        "old_uuid": existing_cfdi.get('uuid'),
        "new_uuid": parsed['uuid']
    }


@router.get("/cfdi", response_model=List[CFDI])
async def list_cfdis(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0)
):
    company_id = await get_active_company_id(request, current_user)
    
    cfdis = await db.cfdis.find(
        {'company_id': company_id},
        {'_id': 0, 'xml_original': 0}
    ).sort('fecha_emision', -1).skip(skip).limit(limit).to_list(limit)
    
    for c in cfdis:
        for field in ['fecha_emision', 'fecha_timbrado', 'created_at']:
            if isinstance(c.get(field), str):
                c[field] = datetime.fromisoformat(c[field])
        
        # Calculate saldo_pendiente for partial payments support
        # Use monto_cobrado for ingresos (sales), monto_pagado for egresos (expenses)
        cfdi_total = c.get('total', 0)
        if c.get('tipo_cfdi') == 'ingreso':
            monto_cubierto = c.get('monto_cobrado', 0) or 0
        else:
            monto_cubierto = c.get('monto_pagado', 0) or 0
        c['saldo_pendiente'] = max(0, cfdi_total - monto_cubierto)
    
    return cfdis

@router.delete("/cfdi/{cfdi_id}")
async def delete_cfdi(cfdi_id: str, request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    
    cfdi = await db.cfdis.find_one({'id': cfdi_id, 'company_id': company_id}, {'_id': 0})
    if not cfdi:
        raise HTTPException(status_code=404, detail="CFDI no encontrado")
    
    await db.cfdis.delete_one({'id': cfdi_id, 'company_id': company_id})
    await audit_log(company_id, 'CFDI', cfdi_id, 'DELETE', current_user['id'])
    
    return {'status': 'success', 'message': 'CFDI eliminado correctamente'}

@router.get("/transactions/template")
async def download_transactions_template():
    """Download Excel template for importing transactions"""
    import io
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    # Headers
    headers = [
        'concepto', 'monto', 'tipo_transaccion', 'fecha_transaccion', 
        'es_proyeccion', 'categoria', 'referencia', 'notas'
    ]
    ws.append(headers)
    
    # Example row
    ws.append([
        'Pago a proveedor', 10000.00, 'egreso', '2026-01-20', 
        'FALSE', 'operativo', 'REF-001', 'Pago factura #123'
    ])
    ws.append([
        'Cobro cliente', 25000.00, 'ingreso', '2026-01-22', 
        'FALSE', 'ventas', 'REF-002', 'Factura #456'
    ])
    
    # Add validation notes
    ws2 = wb.create_sheet("Instrucciones")
    ws2.append(["Campo", "Descripción", "Valores Válidos"])
    ws2.append(["concepto", "Descripción de la transacción", "Texto libre"])
    ws2.append(["monto", "Monto de la transacción", "Número positivo"])
    ws2.append(["tipo_transaccion", "Tipo de movimiento", "ingreso, egreso"])
    ws2.append(["fecha_transaccion", "Fecha del movimiento", "YYYY-MM-DD"])
    ws2.append(["es_proyeccion", "Es proyección futura?", "TRUE, FALSE"])
    ws2.append(["categoria", "Categoría del gasto/ingreso", "Texto libre"])
    ws2.append(["referencia", "Referencia o número de documento", "Texto libre"])
    ws2.append(["notas", "Notas adicionales", "Texto libre"])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_transacciones.xlsx"}
    )


@router.post("/bank-transactions", response_model=BankTransaction)
async def create_bank_transaction(transaction_data: BankTransactionCreate, request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    account = await db.bank_accounts.find_one({'id': transaction_data.bank_account_id, 'company_id': company_id}, {'_id': 0})
    if not account:
        raise HTTPException(status_code=404, detail="Cuenta bancaria no encontrada")
    
    # Use account's currency if not specified in transaction
    txn_data = transaction_data.model_dump()
    if not txn_data.get('moneda'):
        txn_data['moneda'] = account.get('moneda', 'MXN')
    
    bank_transaction = BankTransaction(company_id=company_id, **txn_data)
    doc = bank_transaction.model_dump()
    for field in ['fecha_movimiento', 'fecha_valor', 'created_at']:
        doc[field] = doc[field].isoformat()
    await db.bank_transactions.insert_one(doc)
    
    await audit_log(bank_transaction.company_id, 'BankTransaction', bank_transaction.id, 'CREATE', current_user['id'])
    return bank_transaction

@router.get("/bank-transactions", response_model=List[BankTransaction])
async def list_bank_transactions(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0)
):
    company_id = await get_active_company_id(request, current_user)
    transactions = await db.bank_transactions.find(
        {'company_id': company_id},
        {'_id': 0}
    ).sort('fecha_movimiento', -1).skip(skip).limit(limit).to_list(limit)
    
    for t in transactions:
        for field in ['fecha_movimiento', 'fecha_valor', 'created_at']:
            if isinstance(t.get(field), str):
                t[field] = datetime.fromisoformat(t[field])
    return transactions

@router.delete("/bank-transactions/{transaction_id}")
async def delete_bank_transaction(transaction_id: str, current_user: Dict = Depends(get_current_user)):
    """Delete a bank transaction"""
    # Check if transaction exists
    txn = await db.bank_transactions.find_one({
        'id': transaction_id, 
        'company_id': current_user['company_id']
    })
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    # Check if it's reconciled
    if txn.get('conciliado'):
        raise HTTPException(status_code=400, detail="No se puede eliminar un movimiento conciliado")
    
    # Delete the transaction
    result = await db.bank_transactions.delete_one({
        'id': transaction_id,
        'company_id': current_user['company_id']
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    await audit_log(current_user['company_id'], 'BankTransaction', transaction_id, 'DELETE', current_user['id'])
    return {"status": "success", "message": "Movimiento eliminado"}

@router.post("/bank-transactions/transfer-account")
async def transfer_transactions_to_account(
    data: dict,
    current_user: Dict = Depends(get_current_user)
):
    """
    Transfer transactions from one account to another with optional currency conversion.
    Supports:
    - Transfer between accounts of same bank or different banks
    - Currency conversion using FX rates when moving between MXN/USD accounts
    - Optional custom FX rate override
    """
    company_id = current_user['company_id']
    
    from_account_id = data.get('from_account_id')
    to_account_id = data.get('to_account_id')
    convert_currency = data.get('convert_currency', True)  # Whether to convert amounts
    custom_fx_rate = data.get('custom_fx_rate')  # Optional: user-provided exchange rate
    transaction_ids = data.get('transaction_ids')  # Optional: specific transactions to transfer
    
    if not from_account_id or not to_account_id:
        raise HTTPException(status_code=400, detail="Se requieren from_account_id y to_account_id")
    
    # Verify both accounts exist and belong to company
    from_account = await db.bank_accounts.find_one({'id': from_account_id, 'company_id': company_id}, {'_id': 0})
    to_account = await db.bank_accounts.find_one({'id': to_account_id, 'company_id': company_id}, {'_id': 0})
    
    if not from_account:
        raise HTTPException(status_code=404, detail="Cuenta origen no encontrada")
    if not to_account:
        raise HTTPException(status_code=404, detail="Cuenta destino no encontrada")
    
    from_currency = from_account.get('moneda', 'MXN')
    to_currency = to_account.get('moneda', 'MXN')
    
    # Get FX rate if currencies are different
    fx_rate = 1.0
    if from_currency != to_currency and convert_currency:
        if custom_fx_rate:
            fx_rate = float(custom_fx_rate)
        else:
            # Get latest FX rate from database
            if from_currency == 'USD' and to_currency == 'MXN':
                rate_doc = await db.fx_rates.find_one(
                    {'company_id': company_id, '$or': [{'moneda_cotizada': 'USD'}, {'moneda_origen': 'USD'}]},
                    {'_id': 0},
                    sort=[('fecha_vigencia', -1)]
                )
                fx_rate = rate_doc.get('tipo_cambio') or rate_doc.get('tasa') or 17.5 if rate_doc else 17.5
            elif from_currency == 'MXN' and to_currency == 'USD':
                rate_doc = await db.fx_rates.find_one(
                    {'company_id': company_id, '$or': [{'moneda_cotizada': 'USD'}, {'moneda_origen': 'USD'}]},
                    {'_id': 0},
                    sort=[('fecha_vigencia', -1)]
                )
                base_rate = rate_doc.get('tipo_cambio') or rate_doc.get('tasa') or 17.5 if rate_doc else 17.5
                fx_rate = 1 / base_rate  # Inverse for MXN to USD
            else:
                # For other currency pairs, try to find direct rate
                rate_doc = await db.fx_rates.find_one(
                    {'company_id': company_id, 'moneda_origen': from_currency, 'moneda_cotizada': to_currency},
                    {'_id': 0},
                    sort=[('fecha_vigencia', -1)]
                )
                fx_rate = rate_doc.get('tasa') or 1.0 if rate_doc else 1.0
    
    # Build query for transactions
    txn_query = {'bank_account_id': from_account_id, 'company_id': company_id}
    if transaction_ids and len(transaction_ids) > 0:
        txn_query['id'] = {'$in': transaction_ids}
    
    # Get transactions to transfer
    transactions = await db.bank_transactions.find(txn_query, {'_id': 0}).to_list(10000)
    
    if len(transactions) == 0:
        return {
            "status": "warning",
            "message": "No se encontraron movimientos para transferir",
            "modified_count": 0
        }
    
    # Transfer each transaction with currency conversion if needed
    transferred_count = 0
    total_original = 0
    total_converted = 0
    
    for txn in transactions:
        original_monto = txn.get('monto', 0)
        total_original += original_monto
        
        # Calculate converted amount
        if from_currency != to_currency and convert_currency:
            converted_monto = round(original_monto * fx_rate, 2)
        else:
            converted_monto = original_monto
        
        total_converted += converted_monto
        
        # Update transaction
        update_data = {
            'bank_account_id': to_account_id,
            'moneda': to_currency
        }
        
        if from_currency != to_currency and convert_currency:
            update_data['monto'] = converted_monto
            update_data['monto_original'] = original_monto
            update_data['moneda_original'] = from_currency
            update_data['tipo_cambio_conversion'] = fx_rate
        
        await db.bank_transactions.update_one(
            {'id': txn['id'], 'company_id': company_id},
            {'$set': update_data}
        )
        transferred_count += 1
    
    await audit_log(company_id, 'BankTransaction', 'bulk_transfer', 'UPDATE', current_user['id'], 
                    {
                        'from': from_account_id, 
                        'to': to_account_id, 
                        'count': transferred_count,
                        'from_currency': from_currency,
                        'to_currency': to_currency,
                        'fx_rate': fx_rate,
                        'convert_currency': convert_currency
                    })
    
    return {
        "status": "success",
        "message": f"Se transfirieron {transferred_count} movimientos de {from_account.get('nombre', '')} a {to_account.get('nombre', '')}",
        "modified_count": transferred_count,
        "from_currency": from_currency,
        "to_currency": to_currency,
        "fx_rate_used": fx_rate if from_currency != to_currency else None,
        "total_original": round(total_original, 2),
        "total_converted": round(total_converted, 2),
        "currency_converted": from_currency != to_currency and convert_currency
    }

@router.put("/bank-transactions/{transaction_id}")
async def update_bank_transaction(transaction_id: str, data: dict, current_user: Dict = Depends(get_current_user)):
    """Update a bank transaction"""
    company_id = current_user['company_id']
    
    # Check if transaction exists
    txn = await db.bank_transactions.find_one({
        'id': transaction_id, 
        'company_id': company_id
    })
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    # Fields that can be updated
    allowed_fields = ['bank_account_id', 'descripcion', 'referencia', 'monto', 'tipo_movimiento', 
                      'fecha_movimiento', 'fecha_valor', 'moneda', 'notas']
    
    update_data = {}
    for field in allowed_fields:
        if field in data:
            update_data[field] = data[field]
    
    # If bank_account_id changed, update moneda from the new account
    if 'bank_account_id' in update_data and update_data['bank_account_id']:
        new_account = await db.bank_accounts.find_one({'id': update_data['bank_account_id'], 'company_id': company_id}, {'_id': 0})
        if new_account:
            update_data['moneda'] = new_account.get('moneda', 'MXN')
    
    if not update_data:
        raise HTTPException(status_code=400, detail="No hay campos para actualizar")
    
    result = await db.bank_transactions.update_one(
        {'id': transaction_id, 'company_id': company_id},
        {'$set': update_data}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="No se pudo actualizar el movimiento")
    
    await audit_log(company_id, 'BankTransaction', transaction_id, 'UPDATE', current_user['id'], txn, update_data)
    
    # Return updated transaction
    updated = await db.bank_transactions.find_one({'id': transaction_id}, {'_id': 0})
    return updated

@router.post("/bank-transactions/check-duplicates")
async def check_duplicate_transactions(
    transactions: List[dict],
    current_user: Dict = Depends(get_current_user)
):
    """Check for duplicate transactions before import"""
    company_id = current_user['company_id']
    duplicates = []
    
    for txn in transactions:
        # Check if a transaction with same date, description and amount exists
        query = {
            'company_id': company_id,
            'monto': txn.get('monto'),
            'descripcion': txn.get('descripcion')
        }
        
        # Parse date if string
        fecha = txn.get('fecha_movimiento')
        if fecha:
            if isinstance(fecha, str):
                try:
                    fecha_dt = datetime.fromisoformat(fecha.replace('Z', '+00:00'))
                    # Check within same day
                    start_of_day = fecha_dt.replace(hour=0, minute=0, second=0, microsecond=0)
                    end_of_day = fecha_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
                    query['fecha_movimiento'] = {'$gte': start_of_day.isoformat(), '$lte': end_of_day.isoformat()}
                except:
                    pass
        
        existing = await db.bank_transactions.find_one(query, {'_id': 0})
        if existing:
            duplicates.append({
                'descripcion': txn.get('descripcion'),
                'monto': txn.get('monto'),
                'fecha': str(fecha)[:10] if fecha else ''
            })
    
    return {'duplicates': duplicates, 'count': len(duplicates)}

# ==================== RECONCILIATIONS ENDPOINTS MOVED TO routes/reconciliations.py ====================
# All reconciliation endpoints are now handled by routes/reconciliations.py:
# - POST /reconciliations
# - GET /reconciliations
# - DELETE /reconciliations/{id}
# - GET /reconciliations/by-cfdi/{cfdi_id}
# - GET /reconciliations/summary
# - POST /reconciliations/mark-without-uuid
# - DELETE /reconciliations/bulk/all

# ===== CONCEPTOS MANUALES DE PROYECCIÓN =====
@router.post("/manual-projections", response_model=ManualProjectionConcept)
async def create_manual_projection(data: ManualProjectionConceptCreate, request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    concept = ManualProjectionConcept(company_id=company_id, **data.model_dump())
    doc = concept.model_dump()
    if doc.get('created_at'):
        doc['created_at'] = doc['created_at'].isoformat()
    await db.manual_projections.insert_one(doc)
    await audit_log(company_id, 'ManualProjection', concept.id, 'CREATE', current_user['id'])
    return concept

@router.get("/manual-projections")
async def list_manual_projections(
    request: Request, 
    current_user: Dict = Depends(get_current_user),
    tipo: Optional[str] = Query(None, description="ingreso o egreso"),
    activo: Optional[bool] = Query(True)
):
    company_id = await get_active_company_id(request, current_user)
    query = {'company_id': company_id}
    if tipo:
        query['tipo'] = tipo
    if activo is not None:
        query['activo'] = activo
    concepts = await db.manual_projections.find(query, {'_id': 0}).sort('created_at', -1).to_list(500)
    return concepts

@router.put("/manual-projections/{concept_id}")
async def update_manual_projection(
    concept_id: str, 
    data: ManualProjectionConceptCreate, 
    request: Request, 
    current_user: Dict = Depends(get_current_user)
):
    company_id = await get_active_company_id(request, current_user)
    existing = await db.manual_projections.find_one({'id': concept_id, 'company_id': company_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Concepto no encontrado")
    
    update_data = data.model_dump()
    await db.manual_projections.update_one(
        {'id': concept_id},
        {'$set': update_data}
    )
    await audit_log(company_id, 'ManualProjection', concept_id, 'UPDATE', current_user['id'])
    updated = await db.manual_projections.find_one({'id': concept_id}, {'_id': 0})
    return updated

@router.delete("/manual-projections/{concept_id}")
async def delete_manual_projection(concept_id: str, request: Request, current_user: Dict = Depends(get_current_user)):
    company_id = await get_active_company_id(request, current_user)
    existing = await db.manual_projections.find_one({'id': concept_id, 'company_id': company_id}, {'_id': 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Concepto no encontrado")
    await db.manual_projections.delete_one({'id': concept_id})
    await audit_log(company_id, 'ManualProjection', concept_id, 'DELETE', current_user['id'])
    return {"message": "Concepto eliminado exitosamente"}

# ==================== PAYMENTS ENDPOINTS PARTIALLY MOVED TO routes/payments.py ====================
# Basic CRUD endpoints moved:
# - POST /payments
# - GET /payments
# - PUT /payments/{id}
# - POST /payments/{id}/complete
# - DELETE /payments/{id}
# - DELETE /payments/bulk/all
#
# Specialized endpoints remain here:
# - GET /payments/{id}/match-candidates
# - POST /payments/{id}/auto-reconcile
# - POST /payments/from-bank-with-cfdi-match
# - GET /payments/summary (uses CFDI-based logic)
# - GET /payments/breakdown


