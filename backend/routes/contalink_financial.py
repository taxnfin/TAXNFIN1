"""
contalink_financial.py
Endpoint para importar estados financieros de Contalink (Excel).
Soporta: Balance General, Estado de Resultados.
Detecta el tipo automáticamente por el nombre de la hoja.
"""

from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query, Request
from fastapi.responses import JSONResponse
import openpyxl
import io
import re
from datetime import datetime
from typing import Optional
from core.auth import get_current_user, get_active_company_id
from core.database import db

router = APIRouter(prefix="/contalink-financial", tags=["Contalink Financial"])

def _parse_balance_xls(ws_xls):
    rows = [[ws_xls.cell_value(r, c) for c in range(ws_xls.ncols)] for r in range(ws_xls.nrows)]
    class _M:
        def iter_rows(self, values_only=True): return iter(rows)
    return parse_balance_general(_M())



# ─── Helpers ─────────────────────────────────────────────────────────────────

def detect_indent_level(text: str) -> int:
    """Cuenta espacios iniciales para determinar nivel jerárquico."""
    if not text:
        return 0
    return (len(text) - len(text.lstrip(' '))) // 2


def clean_value(val) -> float:
    """Convierte valor de celda a float limpio, ignorando floating point noise."""
    if val is None:
        return 0.0
    try:
        f = float(val)
        # Ignorar valores de ruido de floating point (< 0.01)
        if abs(f) < 0.01:
            return 0.0
        return round(f, 2)
    except (TypeError, ValueError):
        return 0.0


def clean_label(text: str) -> str:
    """Limpia etiqueta removiendo espacios extras."""
    if not text:
        return ''
    return text.strip()


def extract_date_from_cell(val) -> Optional[str]:
    """Extrae fecha de celda en formato DD/MM/YYYY → YYYY-MM-DD."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val)
    # Buscar patrón DD/MM/YYYY
    match = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if match:
        d, m, y = match.groups()
        return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return None


# ─── Parser Balance General ──────────────────────────────────────────────────

def parse_balance_general(ws) -> dict:
    """
    Parsea la hoja 'Balance General' de Contalink.
    Estructura: 5 columnas, Activo en cols A/B, Pasivo+Capital en cols D/E.
    """
    rows = list(ws.iter_rows(values_only=True))

    # Metadata (filas 1-3)
    empresa   = clean_label(str(rows[0][1])) if rows[0][1] else ''
    rfc_fecha = rows[1] if len(rows) > 1 else []
    rfc       = clean_label(str(rfc_fecha[1])) if len(rfc_fecha) > 1 and rfc_fecha[1] else ''
    fecha_raw = rfc_fecha[4] if len(rfc_fecha) > 4 else None
    fecha     = extract_date_from_cell(fecha_raw)
    moneda    = 'MXN'

    activo_items   = []
    pasivo_items   = []
    capital_items  = []

    current_section_derecha = 'pasivo'  # empieza en PASIVO, luego CAPITAL

    total_activo  = 0.0
    total_pasivo  = 0.0
    total_capital = 0.0

    for row in rows[7:]:  # Skip primeras 7 filas (metadata + encabezados + ACTIVO/PASIVO header)
        col_a, col_b, _, col_d, col_e = (row + (None,)*5)[:5]

        label_izq = clean_label(str(col_a)) if col_a else ''
        val_izq   = clean_value(col_b)
        label_der = clean_label(str(col_d)) if col_d else ''
        val_der   = clean_value(col_e)

        # Detectar sección derecha
        if label_der.upper() == 'CAPITAL':
            current_section_derecha = 'capital'

        # Procesar columna izquierda (ACTIVO)
        if label_izq and label_izq.upper() not in ('ACTIVO', 'TIPO MONEDA: MXN'):
            if label_izq.startswith('Total'):
                total_activo = clean_value(col_b)
            else:
                level = detect_indent_level(str(col_a) if col_a else '')
                activo_items.append({
                    'label': label_izq,
                    'value': val_izq,
                    'level': level,
                })

        # Procesar columna derecha (PASIVO / CAPITAL)
        if label_der and label_der.upper() not in ('PASIVO', 'CAPITAL'):
            if 'Total' in label_der:
                # Ignorar "Total de Pasivo + Capital" (gran total del lado derecho)
                is_grand_total = 'Pasivo' in label_der and 'Capital' in label_der
                if not is_grand_total:
                    if current_section_derecha == 'capital':
                        total_capital = val_der
                    elif 'Pasivo' in label_der:
                        total_pasivo = val_der
            else:
                level = detect_indent_level(str(col_d) if col_d else '')
                item = {
                    'label': label_der,
                    'value': val_der,
                    'level': level,
                }
                if current_section_derecha == 'capital':
                    capital_items.append(item)
                else:
                    pasivo_items.append(item)

    # Calcular totales si no se encontraron en el archivo
    if total_activo == 0:
        total_activo = next(
            (i['value'] for i in activo_items if i['level'] == 0), 0.0
        )
    if total_pasivo == 0:
        # Sum level-1 pasivo sections (Corto + Largo plazo)
        total_pasivo = sum(i['value'] for i in pasivo_items if i['level'] == 1)
    if total_pasivo == 0:
        total_pasivo = next(
            (i['value'] for i in pasivo_items if i['level'] == 0), 0.0
        )

    # Resumen ejecutivo de secciones principales (nivel 1)
    def get_section_summary(items):
        return [
            {'label': i['label'], 'value': i['value']}
            for i in items if i['level'] == 1
        ]

    return {
        'tipo':    'balance_general',
        'empresa': empresa,
        'rfc':     rfc,
        'fecha':   fecha,
        'moneda':  moneda,
        'resumen': {
            'total_activo':         total_activo,
            'total_pasivo':         total_pasivo,
            'total_capital':        total_capital if total_capital != 0
                                    else round(total_activo - total_pasivo, 2),
            'activo_circulante':    next(
                (i['value'] for i in activo_items
                 if 'corto' in i['label'].lower() or 'circulante' in i['label'].lower()), 0.0),
            'activo_fijo':          next(
                (i['value'] for i in activo_items
                 if 'largo' in i['label'].lower() or 'fijo' in i['label'].lower()), 0.0),
            'pasivo_corto_plazo':   next(
                (i['value'] for i in pasivo_items
                 if 'corto' in i['label'].lower()), 0.0),
            'pasivo_largo_plazo':   next(
                (i['value'] for i in pasivo_items
                 if 'largo' in i['label'].lower()), 0.0),
        },
        'activo':  {
            'total':    total_activo,
            'secciones': get_section_summary(activo_items),
            'detalle':  activo_items,
        },
        'pasivo':  {
            'total':    total_pasivo,
            'secciones': get_section_summary(pasivo_items),
            'detalle':  pasivo_items,
        },
        'capital': {
            'total':    total_capital,
            'secciones': get_section_summary(capital_items),
            'detalle':  capital_items,
        },
        # KPIs automáticos
        'kpis': {
            'razon_circulante':     round(
                next((i['value'] for i in activo_items if 'corto' in i['label'].lower()), 0) /
                max(next((i['value'] for i in pasivo_items if 'corto' in i['label'].lower()), 1), 1),
                2),
            'deuda_capital':        round(
                total_pasivo / max(total_capital if total_capital > 0 else
                                   (total_activo - total_pasivo), 1),
                2),
            'solidez':              round(
                total_activo / max(total_pasivo, 1), 2),
        }
    }


# ─── Parser Estado de Resultados ─────────────────────────────────────────────
# Formato real Contalink (.xls / .xlsx):
#   Col A (0): Concepto
#   Col D (3): Monto detalle (float)
#   Col F (5): Subtotal de sección (string con formato " 1,622,593.79")
#   Col H (7): Total final
#   Filas 1-4: metadata (empresa, título, período, centro de costo)

def _parse_subtotal_str(val) -> float:
    """Convierte string con formato ' 1,622,593.79' a float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return clean_value(val)
    s = str(val).strip().replace(',', '').replace(' ', '')
    try:
        f = float(s)
        return round(f, 2) if abs(f) >= 0.01 else 0.0
    except ValueError:
        return 0.0


def _extract_periodo(text: str):
    """Extrae fecha fin de string 'DE 1 DE ENERO DE 2026 AL 31 DE ENERO DE 2026'."""
    if not text:
        return None
    meses = {
        'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
        'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12,
    }
    t = text.lower()
    # buscar "al DD DE MES DE YYYY"
    match = re.search(r'al\s+(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})', t)
    if match:
        d, m_str, y = match.groups()
        m = meses.get(m_str, 1)
        return f"{y}-{str(m).zfill(2)}-{d.zfill(2)}"
    # fallback: buscar cualquier año
    match = re.search(r'(\d{4})', text)
    return match.group(1) + '-01-01' if match else None


def parse_estado_resultados(ws, is_xls=False) -> dict:
    """
    Parsea Estado de Resultados de Contalink.
    Soporta .xlsx (openpyxl) y .xls (xlrd sheet object con interface unificada).
    """
    # Unificar interface: obtener lista de listas
    if is_xls:
        rows = []
        for i in range(ws.nrows):
            rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
    else:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

    empresa = ''
    fecha   = None
    moneda  = 'MXN'
    items   = []  # {'label', 'value', 'subtotal', 'is_total', 'level'}

    # ── Metadata (primeras 4 filas) ──────────────────────────────────────────
    for row in rows[:4]:
        for cell in row:
            if not cell or not isinstance(cell, str):
                continue
            c = cell.strip()
            if not empresa and any(k in c.upper() for k in ('SA DE CV', 'S.A. DE C.V', 'SAPI', 'SC DE RL')):
                empresa = c
            d = _extract_periodo(c)
            if d:
                fecha = d

    # ── Detalle (fila 5 en adelante) ─────────────────────────────────────────
    for row in rows[4:]:
        # Pad row to at least 9 columns
        row = (row + [None] * 9)[:9]
        col_a, col_b, _, col_d, _, col_f, col_g, col_h, _ = row

        # col_a = sección principal, col_b = subcuenta
        label    = clean_label(str(col_a)) if col_a else ''
        label_b  = clean_label(str(col_b)) if col_b else ''  # subcuenta (ej: Depreciación contable)
        detalle  = clean_value(col_d)
        # col_g (índice 6) es el valor del PERIODO en el ER de Contalink
        subtotal = _parse_subtotal_str(col_g) or _parse_subtotal_str(col_f)
        total    = _parse_subtotal_str(col_h) or subtotal

        if not label and subtotal == 0 and total == 0:
            continue

        is_total_line = label.startswith('=') or label.upper().startswith('TOTAL')

        items.append({
            'label':    label,
            'label_b':  label_b,  # subcuenta para buscar depreciación etc.
            'detalle':  detalle,
            'subtotal': subtotal,
            'total':    total,
            'is_total': is_total_line,
        })

    # ── Extraer KPIs clave por label ─────────────────────────────────────────
    def find_subtotal(*keywords):
        """Busca subtotal en la fila que contenga todos los keywords (busca en label y label_b)."""
        for item in items:
            l  = item['label'].upper()
            lb = item.get('label_b', '').upper()
            if all(k.upper() in l for k in keywords) or all(k.upper() in lb for k in keywords):
                v = item['subtotal'] or item['total'] or item['detalle']
                if v:
                    return abs(v)
        return 0.0

    def find_total_seccion(seccion_label):
        """Busca el total de una sección buscando fila 'Total' después de la sección."""
        in_seccion = False
        for item in items:
            l = item['label'].upper()
            if seccion_label.upper() in l:
                in_seccion = True
            if in_seccion and l == 'TOTAL':
                v = item['subtotal'] or item['total']
                if v: return abs(v)
        return 0.0

    ventas_brutas   = find_subtotal('VENTAS', 'GRAVAM') or find_subtotal('INGRESO') or find_total_seccion('INGRESOS')
    devoluciones    = find_subtotal('DEVOLUCION') or find_subtotal('DESCUENTO')
    ventas_netas    = find_subtotal('VENTAS NETAS') or find_subtotal('NETAS')
    costo_ventas    = find_subtotal('COSTO') or find_total_seccion('COSTOS')
    utilidad_bruta  = find_subtotal('UTILIDAD BRUTA') or find_subtotal('BRUTA')
    gastos_admin    = find_subtotal('ADMINISTR')
    gastos_venta    = find_subtotal('GASTO DE VENTA') or find_subtotal('GASTOS DE VENTA')
    ebitda          = find_subtotal('EBITDA')
    depreciacion    = find_subtotal('DEPRECIAC')
    gastos_fin      = find_subtotal('FINANCIERO') or find_subtotal('INTERES') or find_total_seccion('RESULTADO INTEGRAL')
    ebita           = find_subtotal('EBITA') or find_subtotal('EBIT')
    utilidad_neta   = find_subtotal('UTILIDAD TOTAL') or find_subtotal('UTILIDAD NETA') or find_subtotal('NETA')
    # Si no hay gastos separados, usar el total de la sección gastos
    if not gastos_admin and not gastos_venta:
        gastos_total = find_total_seccion('GASTOS')
        gastos_admin = gastos_total

    # Si no hay ventas_netas calcularlo
    if ventas_netas == 0 and ventas_brutas:
        ventas_netas = round(ventas_brutas - devoluciones, 2)

    # Ingresos = ventas_netas si existen, sino ventas_brutas
    ingresos = ventas_netas or ventas_brutas

    # Margen final disponible
    resultado_final = utilidad_neta or ebita or ebitda

    return {
        'tipo':    'estado_resultados',
        'empresa': empresa,
        'rfc':     '',
        'fecha':   fecha,
        'moneda':  moneda,
        'resumen': {
            'ventas_brutas':   ventas_brutas,
            'devoluciones':    devoluciones,
            'ventas_netas':    ventas_netas,
            'costo_ventas':    costo_ventas,
            'utilidad_bruta':  utilidad_bruta,
            'gastos_admin':    gastos_admin,
            'gastos_venta':    gastos_venta,
            'ebitda':          ebitda,
            'depreciacion':    depreciacion,
            'gastos_fin':      gastos_fin,
            'ebita':           ebita,
            'utilidad_neta':   utilidad_neta,
            # Márgenes sobre ventas netas
            'margen_bruto':    round(utilidad_bruta / max(ingresos, 1) * 100, 2) if ingresos else 0,
            'margen_ebitda':   round(ebitda         / max(ingresos, 1) * 100, 2) if ingresos else 0,
            'margen_neto':     round(resultado_final / max(ingresos, 1) * 100, 2) if ingresos else 0,
        },
        'detalle': [
            {'label': i['label'], 'value': i['subtotal'] or i['detalle'], 'is_total': i['is_total']}
            for i in items if i['label']
        ],
    }


# ─── Parser ER por Centro de Costo ───────────────────────────────────────────
# Formato: hoja 'DATOS', 19 cols
#   Col A: sección, Col B: subtítulo, Col C: mayor, Col D: subcuenta
#   Cols E-R: un centro de costo por columna
#   Col S (idx 18): TOTAL consolidado

def parse_er_centro_costo(ws, is_xls=False) -> dict:
    if is_xls:
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
    else:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

    empresa = clean_label(str(rows[0][0])) if rows[0][0] else ''
    fecha   = _extract_periodo(str(rows[2][0])) if len(rows) > 2 else None

    # Centros de costo (fila 4, índice 3): cols 4-17, ignorar vacíos
    header_row = (rows[3] + [None]*20)[:19]
    centros = []
    centro_cols = []
    for idx in range(4, 18):
        h = clean_label(str(header_row[idx])) if header_row[idx] else ''
        if h:
            centros.append(h)
            centro_cols.append(idx)
    total_col = 18  # siempre col 19 (índice 18)

    def get_total(row):
        v = row[total_col] if len(row) > total_col else 0
        return clean_value(v)

    def get_centros(row):
        result = {}
        for ci, col in enumerate(centro_cols):
            v = clean_value(row[col]) if len(row) > col else 0.0
            if v != 0.0:
                result[centros[ci]] = v
        return result

    # Buscar filas clave
    totales = {}
    key_labels = {
        'TOTAL INGRESOS':   'ingresos',
        'TOTAL ACREEDORAS': 'ingresos',  # alias
        'TOTAL COSTOS':     'costo_ventas',
        'TOTAL GASTOS':     'gastos',
        'TOTAL RESULTADO INTEGRAL': 'gastos_fin',
        'TOTAL DEUDORAS':   'total_egresos',
        'UTILIDAD NETA':    'utilidad_neta',
    }

    for row in rows[4:]:
        row = (row + [None]*20)[:19]
        col_a = clean_label(str(row[0])) if row[0] else ''
        col_b = clean_label(str(row[1])) if row[1] else ''
        label = col_a or col_b

        for key, field in key_labels.items():
            if key in label.upper() and field not in totales:
                totales[field] = {
                    'label':      label,
                    'total':      get_total(row),
                    'por_centro': get_centros(row),
                }

    ingresos     = totales.get('ingresos',    {}).get('total', 0)
    costo        = totales.get('costo_ventas', {}).get('total', 0)
    gastos       = totales.get('gastos',       {}).get('total', 0)
    gastos_fin   = totales.get('gastos_fin',   {}).get('total', 0)
    util_neta_t  = totales.get('utilidad_neta',{}).get('total', 0)
    util_por_c   = totales.get('utilidad_neta',{}).get('por_centro', {})
    ing_por_c    = totales.get('ingresos',     {}).get('por_centro', {})

    # Ranking centros por utilidad
    ranking = sorted(
        [{'centro': k, 'utilidad': v, 'ingresos': ing_por_c.get(k, 0)}
         for k, v in util_por_c.items() if k and '~' not in k],
        key=lambda x: x['utilidad'], reverse=True
    )

    return {
        'tipo':     'er_centro_costo',
        'empresa':  empresa,
        'fecha':    fecha,
        'moneda':   'MXN',
        'centros':  centros,
        'resumen':  {
            'ingresos':      ingresos,
            'costo_ventas':  costo,
            'utilidad_bruta': round(ingresos - costo, 2),
            'gastos':        gastos,
            'gastos_fin':    gastos_fin,
            'utilidad_neta': util_neta_t,
            'margen_bruto':  round((ingresos - costo) / max(ingresos,1) * 100, 2),
            'margen_neto':   round(util_neta_t / max(ingresos,1) * 100, 2),
        },
        'por_centro': {
            'ingresos':      ing_por_c,
            'utilidad_neta': util_por_c,
        },
        'ranking_centros': ranking,
        'totales_detalle': totales,
    }


# ─── Endpoint principal ───────────────────────────────────────────────────────

@router.post("/import")
async def import_estado_financiero(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Recibe un Excel de Contalink (.xlsx o .xls) y retorna datos parseados.
    Detecta automáticamente: Balance General o Estado de Resultados.
    """
    import xlrd as _xlrd

    fname = (file.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx, .xls)")

    contents = await file.read()
    is_xls = fname.endswith('.xls')

    try:
        if is_xls:
            wb_xls = _xlrd.open_workbook(file_contents=contents)
            sheet_names = [s.lower() for s in wb_xls.sheet_names()]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]
    except Exception as e:
        raise HTTPException(400, f"Error leyendo el archivo Excel: {str(e)}")

    if is_xls:
        def get_ws(keyword):
            name = next((s for s in wb_xls.sheet_names() if keyword in s.lower()), None)
            return wb_xls.sheet_by_name(name) if name else wb_xls.sheet_by_index(0)
        get_flat = lambda ws: ' '.join(
            str(ws.cell_value(r, c))
            for r in range(min(ws.nrows, 10)) for c in range(ws.ncols)
        ).upper()
    else:
        def get_ws(keyword):
            name = next((s for s in wb.sheetnames if keyword in s.lower()), None)
            return wb[name] if name else wb[wb.sheetnames[0]]
        get_flat = lambda ws: ' '.join(
            str(v) for row in ws.iter_rows(values_only=True) for v in row if v
        )[:1000].upper()

    if any('balance' in s for s in sheet_names):
        ws = get_ws('balance')
        result = parse_balance_general(ws) if not is_xls else _parse_balance_xls(ws)
    elif any('datos' in s for s in sheet_names) or any('centro' in s for s in sheet_names):
        # ER por Centro de Costo — hoja llamada 'DATOS'
        ws = get_ws('datos') if any('datos' in s for s in sheet_names) else get_ws('centro')
        result = parse_er_centro_costo(ws, is_xls=is_xls)
    elif any('resultado' in s or 'reportdemo' in s for s in sheet_names):
        ws_name = next((s for s in (wb_xls.sheet_names() if is_xls else wb.sheetnames) if 'resultado' in s.lower() or 'reportdemo' in s.lower()), None)
        ws = wb_xls.sheet_by_name(ws_name) if is_xls else wb[ws_name]
        result = parse_estado_resultados(ws, is_xls=is_xls)
    else:
        ws = get_ws('')
        flat = get_flat(ws)
        if 'BALANCE' in flat or 'ACTIVO' in flat:
            result = parse_balance_general(ws) if not is_xls else _parse_balance_xls(ws)
        elif 'CENTRO DE COSTO' in flat or 'SUBTITULO' in flat:
            result = parse_er_centro_costo(ws, is_xls=is_xls)
        elif 'RESULTADO' in flat or 'INGRESO' in flat or 'VENTA' in flat:
            result = parse_estado_resultados(ws, is_xls=is_xls)
        else:
            raise HTTPException(400,
                "No se pudo detectar el tipo. Exporta desde Contalink: "
                "Balance General, Estado de Resultados o ER por Centro de Costo en Excel."
            )

    return JSONResponse(content={'success': True, 'data': result, 'archivo': file.filename})


# ─── Endpoint: guardar en DB (opcional para historial) ───────────────────────

@router.post("/save")
async def save_estado_financiero(
    payload: dict,
    current_user: dict = Depends(get_current_user),
):
    """
    Guarda el estado financiero parseado en MongoDB para historial.
    """
    collection = db['estados_financieros_contalink']

    doc = {
        **payload,
        'empresa_id':   get_active_company_id(current_user),
        'usuario_id':   current_user.get('id'),
        'importado_en': datetime.utcnow().isoformat(),
    }

    # Evitar duplicados por tipo + fecha
    existing = await collection.find_one({
        'empresa_id': doc['empresa_id'],
        'tipo':       doc.get('tipo'),
        'fecha':      doc.get('fecha'),
    })

    if existing:
        await collection.replace_one({'_id': existing['_id']}, doc)
        return {'success': True, 'action': 'updated', 'id': str(existing['_id'])}
    else:
        result = await collection.insert_one(doc)
        return {'success': True, 'action': 'created', 'id': str(result.inserted_id)}


@router.get("/history")
async def get_history(current_user: dict = Depends(get_current_user)):
    """Retorna historial de estados financieros importados."""
    collection = db['estados_financieros_contalink']
    docs = await collection.find(
        {'empresa_id': current_user.get('empresa_id')},
        {'_id': 0, 'tipo': 1, 'fecha': 1, 'empresa': 1, 'importado_en': 1, 'resumen': 1}
    ).sort('importado_en', -1).limit(24).to_list(24)
    return {'success': True, 'data': docs}


# ─── Endpoints para alimentar Financial Statements (Métricas + Reporte Board) ──

@router.post("/upload-to-metrics")
async def upload_contalink_to_metrics(
    request: Request,
    file: UploadFile = File(...),
    periodo: str = Query(..., description="Período YYYY-MM, ej: 2026-01"),
    current_user: dict = Depends(get_current_user),
):
    """
    Procesa un Excel de Contalink y lo guarda en financial_statements
    para que lo lean Métricas Financieras y Reporte Board.
    Acepta: Balance General, Estado de Resultados, ER por Centro de Costo.
    """
    import xlrd as _xlrd
    from core.auth import get_active_company_id

    fname = (file.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Solo se aceptan archivos Excel (.xlsx, .xls)")

    contents = await file.read()
    is_xls   = fname.endswith('.xls')
    company_id = str(request.headers.get('X-Company-ID') or request.headers.get('x-company-id') or current_user.get('company_id', ''))

    try:
        if is_xls:
            wb_xls = _xlrd.open_workbook(file_contents=contents)
            sheet_names = [s.lower() for s in wb_xls.sheet_names()]
        else:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet_names = [s.lower() for s in wb.sheetnames]
    except Exception as e:
        raise HTTPException(400, f"Error leyendo archivo: {str(e)}")

    # Detectar tipo
    if any('balance' in s for s in sheet_names):
        # Balance General → guardar como balance_general
        if is_xls:
            ws = wb_xls.sheet_by_index(0)
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        else:
            ws_obj = wb[next(s for s in wb.sheetnames if 'balance' in s.lower())]
            rows = [list(r) for r in ws_obj.iter_rows(values_only=True)]

        # Parse con nuestro parser existente
        parsed = _parse_balance_xls(ws) if is_xls else parse_balance_general(ws_obj)

        # Convertir al formato de financial_statements
        res = parsed.get('resumen', {})
        activo  = parsed.get('activo', {})
        pasivo  = parsed.get('pasivo', {})
        capital = parsed.get('capital', {})

        # Extraer valores del detalle
        def find_val(items, *keywords):
            for item in (items or []):
                label = item.get('label', '').lower()
                if all(k in label for k in keywords):
                    v = item.get('value', 0)
                    if v and abs(v) > 0.01:
                        return abs(v)
            return 0

        activo_det  = activo.get('detalle', [])
        pasivo_det  = pasivo.get('detalle', [])
        capital_det = capital.get('detalle', [])

        balance_doc = {
            'activo_circulante':         res.get('activo_circulante', 0),
            'efectivo':                  find_val(activo_det, 'banco') or find_val(activo_det, 'efectivo') or find_val(activo_det, 'caja'),
            'cuentas_por_cobrar':        find_val(activo_det, 'cliente'),
            'inventarios':               find_val(activo_det, 'inventario'),
            'otros_activos_circulantes': 0,
            'activo_fijo':               res.get('activo_fijo', 0),
            'activo_total':              res.get('total_activo', 0),
            'pasivo_circulante':         res.get('pasivo_corto_plazo', 0),
            'cuentas_por_pagar':         find_val(pasivo_det, 'proveedor'),
            'deuda_corto_plazo':         0,
            'otros_pasivos_circulantes': 0,
            'pasivo_largo_plazo':        res.get('pasivo_largo_plazo', 0),
            'deuda_largo_plazo':         0,
            'pasivo_total':              res.get('total_pasivo', 0),
            'capital_social':            find_val(capital_det, 'capital'),
            'utilidades_retenidas':      find_val(capital_det, 'resultado') or find_val(capital_det, 'utilidad'),
            'capital_contable':          res.get('total_capital', 0),
            'raw_data':                  [],
        }

        doc = {
            'company_id':  company_id,
            'tipo':        'balance_general',
            'periodo':     periodo,
            'año':         int(periodo.split('-')[0]),
            'mes':         int(periodo.split('-')[1]),
            'fuente':      'contalink',
            'archivo':     file.filename,
            'datos':        balance_doc,
            'created_at':  datetime.utcnow().isoformat(),
            'updated_at':  datetime.utcnow().isoformat(),
        }

        await db.financial_statements.delete_many({
            'company_id': company_id, 'tipo': 'balance_general', 'periodo': periodo
        })
        await db.financial_statements.insert_one(doc)

        return JSONResponse({'success': True, 'tipo': 'balance_general', 'periodo': periodo, 'action': 'guardado',
                             'activo_total': balance_doc['activo_total'], 'pasivo_total': balance_doc['pasivo_total']})

    elif any('resultado' in s or 'reportdemo' in s or 'datos' in s for s in sheet_names):
        # Estado de Resultados → guardar como estado_resultados
        if is_xls:
            ws = wb_xls.sheet_by_index(0)
        else:
            ws_obj = wb[wb.sheetnames[0]]

        parsed = parse_estado_resultados(ws, is_xls=is_xls) if is_xls else parse_estado_resultados(ws_obj, is_xls=False)
        res = parsed.get('resumen', {})

        # Si es ER por centro de costo, usar parse_er_centro_costo
        if any('datos' in s for s in sheet_names):
            parsed = parse_er_centro_costo(ws if is_xls else ws_obj, is_xls=is_xls)
            res = parsed.get('resumen', {})

        ingresos     = res.get('ventas_netas', 0) or res.get('ventas_brutas', 0) or res.get('ingresos', 0)
        costo_ventas = res.get('costo_ventas', 0)
        gastos_admin = res.get('gastos_admin', 0)
        gastos_venta = res.get('gastos_venta', 0)
        gastos_fin   = res.get('gastos_fin', 0)
        depreciacion = res.get('depreciacion', 0)
        util_neta    = res.get('utilidad_neta', 0) or res.get('ebita', 0)

        # Calcular derivados si no vienen explícitos
        util_bruta = res.get('utilidad_bruta', 0)
        if not util_bruta and ingresos and costo_ventas:
            util_bruta = round(ingresos - costo_ventas, 2)

        gastos_op = round((gastos_admin or 0) + (gastos_venta or 0), 2)

        util_op = res.get('utilidad_operativa', 0)
        if not util_op and util_bruta:
            util_op = round(util_bruta - gastos_op, 2)

        ebitda = res.get('ebitda', 0)
        if not ebitda and util_op:
            ebitda = round(util_op + (depreciacion or 0), 2)

        income_doc = {
            'ingresos':                 ingresos,
            'costo_ventas':             costo_ventas,
            'utilidad_bruta':           util_bruta,
            'gastos_venta':             gastos_venta,
            'gastos_administracion':    gastos_admin,
            'gastos_generales':         gastos_op,
            'utilidad_operativa':       util_op,
            'otros_ingresos':           0,
            'gastos_financieros':       gastos_fin,
            'otros_gastos':             0,
            'utilidad_antes_impuestos': util_neta,
            'impuestos':                0,
            'utilidad_neta':            util_neta,
            'depreciacion':             depreciacion,
            'amortizacion':             0,
            'intereses':                gastos_fin,
            'ebitda':                   ebitda,
            'raw_data':                 [],
        }

        doc = {
            'company_id':  company_id,
            'tipo':        'estado_resultados',
            'periodo':     periodo,
            'año':         int(periodo.split('-')[0]),
            'mes':         int(periodo.split('-')[1]),
            'fuente':      'contalink',
            'archivo':     file.filename,
            'datos':        income_doc,
            'created_at':  datetime.utcnow().isoformat(),
            'updated_at':  datetime.utcnow().isoformat(),
        }

        await db.financial_statements.delete_many({
            'company_id': company_id, 'tipo': 'estado_resultados', 'periodo': periodo
        })
        await db.financial_statements.insert_one(doc)

        return JSONResponse({'success': True, 'tipo': 'estado_resultados', 'periodo': periodo, 'action': 'guardado',
                             'ingresos': income_doc['ingresos'], 'utilidad_neta': income_doc['utilidad_neta']})

    else:
        raise HTTPException(400, "No se pudo detectar el tipo de reporte. Sube Balance General o Estado de Resultados de Contalink.")


@router.get("/debug/{periodo}")
async def debug_financial_statement(
    request: Request,
    periodo: str,
    current_user: dict = Depends(get_current_user),
):
    """Debug endpoint to see what was saved in financial_statements collection."""
    company_id = str(request.headers.get('X-Company-ID') or request.headers.get('x-company-id') or current_user.get('company_id', ''))
    
    docs = await db.financial_statements.find(
        {'company_id': company_id, 'periodo': periodo},
        {'_id': 0, 'tipo': 1, 'periodo': 1, 'fuente': 1, 'data': 1}
    ).to_list(10)
    
    return {
        'company_id': company_id,
        'periodo': periodo,
        'count': len(docs),
        'docs': docs,
    }
