"""Bank Transactions routes - Bank statement management"""
from fastapi import APIRouter, Depends, HTTPException, Request, Query
from typing import Dict, List, Optional
from datetime import datetime, timezone
import logging

from core.database import db
from core.auth import get_current_user, get_active_company_id
from models.bank import BankTransaction, BankTransactionCreate
from services.audit import audit_log

router = APIRouter(prefix="/bank-transactions")
logger = logging.getLogger(__name__)


@router.get("/template")
async def download_bank_statement_template():
    """Download Excel template for importing bank statements (NO AUTH REQUIRED)"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    
    headers = [
        'fecha_movimiento', 'fecha_valor', 'descripcion', 'referencia',
        'monto', 'tipo_movimiento', 'saldo', 'categoria', 'notas'
    ]
    
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    example_data = [
        ['2026-01-15', '2026-01-15', 'TRANSFERENCIA SPEI CLIENTE ABC', 'REF123456', 50000.00, 'credito', 150000.00, 'Ventas', 'Pago factura 001'],
        ['2026-01-16', '2026-01-16', 'PAGO NOMINA ENERO', 'NOM202601', -25000.00, 'debito', 125000.00, 'Nómina', 'Quincena 1'],
        ['2026-01-17', '2026-01-17', 'COMISION BANCARIA', 'COM0117', -150.00, 'debito', 124850.00, 'Comisiones Bancarias', ''],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
    
    ws_inst = wb.create_sheet("Instrucciones")
    instructions = [
        ["PLANTILLA PARA IMPORTAR ESTADO DE CUENTA"],
        [""],
        ["Campos requeridos:"],
        ["- fecha_movimiento: Fecha del movimiento (formato: YYYY-MM-DD)"],
        ["- descripcion: Descripción del movimiento"],
        ["- monto: Monto del movimiento (positivo=abono, negativo=cargo)"],
        ["- tipo_movimiento: 'credito' para depósitos, 'debito' para retiros"],
        [""],
        ["Campos opcionales:"],
        ["- fecha_valor: Fecha valor (por defecto igual a fecha_movimiento)"],
        ["- referencia: Número de referencia bancaria"],
        ["- saldo: Saldo después del movimiento"],
        ["- categoria: Categoría del movimiento"],
        ["- notas: Notas adicionales"],
        [""],
        ["IMPORTANTE:"],
        ["1. No modifique los nombres de las columnas"],
        ["2. Los montos negativos se consideran retiros"],
        ["3. Seleccione la cuenta bancaria al importar"],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        cell = ws_inst.cell(row=row_idx, column=1, value=row_data[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "Campos" in str(row_data[0]) or "IMPORTANTE" in str(row_data[0]):
            cell.font = Font(bold=True)
    
    ws_inst.column_dimensions['A'].width = 60
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_estado_cuenta.xlsx"}
    )


@router.post("", response_model=BankTransaction)
async def create_bank_transaction(
    transaction_data: BankTransactionCreate, 
    request: Request, 
    current_user: Dict = Depends(get_current_user)
):
    """Create a new bank transaction"""
    company_id = await get_active_company_id(request, current_user)
    account = await db.bank_accounts.find_one({'id': transaction_data.bank_account_id, 'company_id': company_id}, {'_id': 0})
    if not account:
        raise HTTPException(status_code=404, detail="Cuenta bancaria no encontrada")
    
    # Use account's currency if not specified in transaction
    txn_data = transaction_data.model_dump()
    if not txn_data.get('moneda'):
        txn_data['moneda'] = account.get('moneda', 'MXN')
    
    bank_transaction = BankTransaction(company_id=company_id, **txn_data)
    doc = bank_transaction.model_dump()
    for field in ['fecha_movimiento', 'fecha_valor', 'created_at']:
        if doc.get(field):
            doc[field] = doc[field].isoformat()
    await db.bank_transactions.insert_one(doc)
    
    await audit_log(bank_transaction.company_id, 'BankTransaction', bank_transaction.id, 'CREATE', current_user['id'])
    return bank_transaction


@router.get("", response_model=List[BankTransaction])
async def list_bank_transactions(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    bank_account_id: Optional[str] = Query(None),
    conciliado: Optional[bool] = Query(None),
    fecha_desde: Optional[str] = Query(None),
    fecha_hasta: Optional[str] = Query(None),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0)
):
    """List bank transactions with optional filters"""
    company_id = await get_active_company_id(request, current_user)
    
    query = {'company_id': company_id}
    if bank_account_id:
        query['bank_account_id'] = bank_account_id
    if conciliado is not None:
        query['conciliado'] = conciliado
    if fecha_desde or fecha_hasta:
        query['fecha_movimiento'] = {}
        if fecha_desde:
            query['fecha_movimiento']['$gte'] = fecha_desde
        if fecha_hasta:
            query['fecha_movimiento']['$lte'] = fecha_hasta + 'T23:59:59'
    
    transactions = await db.bank_transactions.find(
        query,
        {'_id': 0}
    ).sort('fecha_movimiento', -1).skip(skip).limit(limit).to_list(limit)
    
    for t in transactions:
        for field in ['fecha_movimiento', 'fecha_valor', 'created_at']:
            if isinstance(t.get(field), str):
                try:
                    t[field] = datetime.fromisoformat(t[field])
                except:
                    pass
    return transactions


@router.get("/{transaction_id}")
async def get_bank_transaction(
    transaction_id: str,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Get a single bank transaction"""
    company_id = await get_active_company_id(request, current_user)
    
    txn = await db.bank_transactions.find_one(
        {'id': transaction_id, 'company_id': company_id},
        {'_id': 0}
    )
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    return txn


@router.put("/{transaction_id}")
async def update_bank_transaction(
    transaction_id: str,
    data: dict,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Update a bank transaction"""
    company_id = await get_active_company_id(request, current_user)
    
    txn = await db.bank_transactions.find_one(
        {'id': transaction_id, 'company_id': company_id},
        {'_id': 0}
    )
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    # Don't allow editing reconciled transactions (except notes)
    if txn.get('conciliado') and not (len(data) == 1 and 'notas' in data):
        raise HTTPException(status_code=400, detail="No se puede editar un movimiento conciliado")
    
    # Convert date fields
    for field in ['fecha_movimiento', 'fecha_valor']:
        if field in data and data[field]:
            if isinstance(data[field], str):
                pass  # Keep as string
            elif hasattr(data[field], 'isoformat'):
                data[field] = data[field].isoformat()
    
    await db.bank_transactions.update_one(
        {'id': transaction_id, 'company_id': company_id},
        {'$set': data}
    )
    
    await audit_log(company_id, 'BankTransaction', transaction_id, 'UPDATE', current_user['id'])
    return {'status': 'success', 'message': 'Movimiento actualizado'}


@router.delete("/{transaction_id}")
async def delete_bank_transaction(
    transaction_id: str, 
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Delete a bank transaction"""
    company_id = await get_active_company_id(request, current_user)
    
    txn = await db.bank_transactions.find_one({'id': transaction_id, 'company_id': company_id}, {'_id': 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    # Check if it's reconciled
    if txn.get('conciliado'):
        raise HTTPException(status_code=400, detail="No se puede eliminar un movimiento conciliado. Cancele primero la conciliación.")
    
    # Delete the transaction
    result = await db.bank_transactions.delete_one({'id': transaction_id, 'company_id': company_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    await audit_log(company_id, 'BankTransaction', transaction_id, 'DELETE', current_user['id'])
    return {"status": "success", "message": "Movimiento eliminado"}


@router.get("/{txn_id}/match-cfdi")
async def get_cfdi_matches_for_transaction(
    txn_id: str,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Find potential CFDI matches for a bank transaction"""
    company_id = await get_active_company_id(request, current_user)
    
    txn = await db.bank_transactions.find_one({'id': txn_id, 'company_id': company_id}, {'_id': 0})
    if not txn:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    
    monto = txn.get('monto', 0)
    tipo = txn.get('tipo_movimiento', '')
    moneda = txn.get('moneda', 'MXN')
    descripcion_original = txn.get('descripcion', '')
    referencia = txn.get('referencia', '')
    descripcion = (descripcion_original + ' ' + referencia).upper()
    
    # Determine CFDI type based on transaction type
    tipo_cfdi = 'ingreso' if tipo == 'credito' else 'egreso'
    
    # Find matching CFDIs
    query = {
        'company_id': company_id,
        'tipo_cfdi': tipo_cfdi if tipo == 'credito' else {'$in': ['egreso', 'nomina']},
        'estado_conciliacion': {'$ne': 'conciliado'},
    }
    
    cfdis = await db.cfdis.find(query, {'_id': 0, 'xml_original': 0}).to_list(1000)
    
    matches = []
    for cfdi in cfdis:
        cfdi_total = cfdi.get('total', 0)
        cfdi_moneda = cfdi.get('moneda', 'MXN')
        cfdi_emisor = cfdi.get('emisor_nombre', '').upper()
        cfdi_receptor = cfdi.get('receptor_nombre', '').upper()
        
        # Calculate pending amount
        if cfdi.get('tipo_cfdi') == 'ingreso':
            monto_cubierto = cfdi.get('monto_cobrado', 0) or 0
        else:
            monto_cubierto = cfdi.get('monto_pagado', 0) or 0
        
        saldo_pendiente = cfdi_total - monto_cubierto
        
        # Skip fully paid CFDIs
        if saldo_pendiente < 0.01:
            continue
        
        # Calculate match score
        score = 0
        match_reasons = []
        
        # Amount match - compare with both total and pending
        diff_pct_pending = abs(saldo_pendiente - monto) / max(saldo_pendiente, 0.01) * 100
        diff_pct_total = abs(cfdi_total - monto) / max(cfdi_total, 0.01) * 100
        diff_pct = min(diff_pct_pending, diff_pct_total)
        
        if diff_pct < 0.1:  # Exact match
            score += 55
            match_reasons.append("Monto exacto")
        elif diff_pct < 1:  # Within 1%
            score += 50
            match_reasons.append("Monto exacto")
        elif diff_pct < 5:  # Within 5%
            score += 35
            match_reasons.append(f"Monto cercano ({diff_pct:.1f}%)")
        elif diff_pct < 10:  # Within 10%
            score += 20
            match_reasons.append(f"Monto aproximado ({diff_pct:.1f}%)")
        else:
            continue  # Skip if amount too different
        
        # Currency match bonus
        if cfdi_moneda == moneda:
            score += 10
            match_reasons.append("Moneda coincide")
        
        # Name matching - improved for truncated bank descriptions
        nombres_buscar = [cfdi_emisor, cfdi_receptor]
        
        for nombre in nombres_buscar:
            if not nombre or len(nombre) < 3:
                continue
            
            nombre_parts = [p for p in nombre.split() if len(p) > 2]
            descripcion_parts = [p for p in descripcion.split() if len(p) > 2]
            
            # Method 1: Count CFDI name parts in description
            matches_count = sum(1 for part in nombre_parts if part in descripcion)
            
            if matches_count >= 3:
                score += 50
                match_reasons.append("Nombre completo")
                break
            elif matches_count >= 2:
                score += 40
                match_reasons.append(f"Nombre parcial ({matches_count} partes)")
                break
            elif matches_count >= 1 and len(nombre_parts) <= 3:
                score += 25
                match_reasons.append("Nombre parcial")
                break
            
            # Method 2: Check if truncated description matches start of CFDI name
            # e.g., "INGENIERIA EN MAQUINARIA Y ENE" matches "INGENIERIA EN MAQUINARIA Y ENERGIA..."
            if descripcion_parts and len(descripcion_parts) >= 2:
                consecutive_matches = 0
                for i, desc_part in enumerate(descripcion_parts):
                    if i < len(nombre_parts) and desc_part in nombre_parts[i]:
                        consecutive_matches += 1
                    elif desc_part in nombre:
                        consecutive_matches += 0.5
                
                if consecutive_matches >= len(descripcion_parts) * 0.7:
                    score += 45
                    match_reasons.append("Nombre truncado coincide")
                    break
            
            # Method 3: Check if description contains start of name (without spaces)
            nombre_sin_espacios = ''.join(nombre.split())[:30]
            desc_sin_espacios = ''.join(descripcion.split())[:30]
            
            if len(desc_sin_espacios) >= 10 and nombre_sin_espacios.startswith(desc_sin_espacios[:15]):
                score += 40
                match_reasons.append("Nombre truncado")
                break
        
        # Only include if score is meaningful
        if score >= 20:
            matches.append({
                'cfdi': cfdi,
                'score': score,
                'match_reasons': match_reasons
            })
    
    # Sort by score
    matches.sort(key=lambda x: x['score'], reverse=True)
    
    # Find best match
    best_match = None
    if matches:
        best = matches[0]
        best_match = {
            **best['cfdi'],
            'score': best['score'],
            'match_reasons': best['match_reasons']
        }
    
    return {
        'matches': matches[:10], 
        'total_candidates': len(matches),
        'best_match': best_match,
        'all_matches': [m['cfdi'] | {'score': m['score'], 'match_reasons': m['match_reasons']} for m in matches[:10]]
    }


@router.post("/check-duplicates")
async def check_duplicate_transactions(
    data: dict,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Check for potential duplicate transactions"""
    company_id = await get_active_company_id(request, current_user)
    
    transactions = data.get('transactions', [])
    duplicates = []
    
    for txn in transactions:
        # Check if similar transaction exists
        fecha = txn.get('fecha_movimiento', '')
        monto = txn.get('monto', 0)
        referencia = txn.get('referencia', '')
        
        query = {
            'company_id': company_id,
            'monto': monto
        }
        
        if fecha:
            query['fecha_movimiento'] = {'$regex': fecha[:10]}
        if referencia:
            query['referencia'] = referencia
        
        existing = await db.bank_transactions.find_one(query, {'_id': 0})
        if existing:
            duplicates.append({
                'new': txn,
                'existing': existing
            })
    
    return {'duplicates': duplicates, 'count': len(duplicates)}


@router.delete("/bulk/all")
async def delete_all_bank_transactions(
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Delete ALL bank transactions for the current company"""
    company_id = await get_active_company_id(request, current_user)
    
    # First delete all reconciliations
    await db.reconciliations.delete_many({'company_id': company_id})
    
    # Delete all payments linked to bank transactions
    await db.payments.delete_many({'company_id': company_id, 'bank_transaction_id': {'$exists': True}})
    
    # Delete all bank transactions
    result = await db.bank_transactions.delete_many({'company_id': company_id})
    
    await audit_log(company_id, 'BankTransaction', 'BULK_DELETE', 'DELETE', current_user['id'],
                    {'count': result.deleted_count})
    
    return {
        'status': 'success',
        'message': f'Se eliminaron {result.deleted_count} movimientos bancarios',
        'deleted_count': result.deleted_count
    }


@router.delete("/bulk-delete")
async def bulk_delete_bank_transactions(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    bank_account_id: Optional[str] = Query(None, description="Filter by bank account"),
    estado_conciliacion: Optional[str] = Query(None, description="Filter by reconciliation status")
):
    """Delete bank transactions with optional filters"""
    company_id = await get_active_company_id(request, current_user)
    
    # Build query filter
    query = {'company_id': company_id}
    
    if bank_account_id:
        query['bank_account_id'] = bank_account_id
    
    if estado_conciliacion:
        query['estado_conciliacion'] = estado_conciliacion
    
    # Get transaction IDs to delete
    transactions_to_delete = await db.bank_transactions.find(query, {'id': 1, '_id': 0}).to_list(50000)
    transaction_ids = [t['id'] for t in transactions_to_delete]
    
    if not transaction_ids:
        return {'status': 'success', 'message': 'No hay movimientos para eliminar', 'deleted': 0}
    
    # Delete associated reconciliations
    await db.reconciliations.delete_many({
        'company_id': company_id,
        'bank_transaction_id': {'$in': transaction_ids}
    })
    
    # Delete bank transactions
    result = await db.bank_transactions.delete_many(query)
    
    await audit_log(company_id, 'BankTransaction', f'bulk_delete_{len(transaction_ids)}', 'BULK_DELETE', current_user['id'])
    
    return {
        'status': 'success',
        'message': f'Se eliminaron {result.deleted_count} movimientos bancarios',
        'deleted': result.deleted_count
    }
