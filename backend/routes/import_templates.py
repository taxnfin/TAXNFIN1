"""Vendor/Customer import templates and bulk import routes"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request, Query
from fastapi.responses import StreamingResponse
from typing import Dict, List, Optional, Any
from datetime import datetime, timezone
import uuid
import io
import openpyxl
from openpyxl import Workbook

from core.database import db
from core.auth import get_current_user
from services.audit import audit_log
from models.enums import UserRole

router = APIRouter()

@router.get("/vendors/template")
async def download_vendors_template():
    """Download Excel template for importing vendors"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Proveedores"
    
    # Headers
    headers = ['nombre', 'rfc', 'email', 'telefono', 'direccion', 'condiciones_pago', 'notas']
    ws.append(headers)
    
    # Example rows
    ws.append(['Proveedor Ejemplo SA de CV', 'PEJ123456ABC', 'contacto@proveedor.com', '5512345678', 'Av. Ejemplo 123, CDMX', '30 días', 'Proveedor de materiales'])
    ws.append(['Servicios Profesionales SA', 'SPS987654XYZ', 'info@servicios.com', '5587654321', 'Calle Servicios 456', '15 días', 'Servicios de consultoría'])
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_proveedores.xlsx"}
    )

@router.get("/customers/template")
async def download_customers_template():
    """Download Excel template for importing customers"""
    from openpyxl import Workbook
    from fastapi.responses import StreamingResponse
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Headers
    headers = ['nombre', 'rfc', 'email', 'telefono', 'direccion', 'limite_credito', 'notas']
    ws.append(headers)
    
    # Example rows
    ws.append(['Cliente Ejemplo SA de CV', 'CEJ123456ABC', 'compras@cliente.com', '5512345678', 'Av. Cliente 123, CDMX', '100000', 'Cliente frecuente'])
    ws.append(['Comercializadora XYZ SA', 'CXY987654XYZ', 'pagos@xyz.com', '5587654321', 'Calle Comercio 456', '50000', 'Nuevo cliente'])
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = cell.font.copy(bold=True)
        ws.column_dimensions[cell.column_letter].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_clientes.xlsx"}
    )

@router.post("/vendors/import")
async def import_vendors(request: Request, file: UploadFile = File(...), current_user: Dict = Depends(get_current_user)):
    """Import vendors from Excel file"""
    from openpyxl import load_workbook
    import io
    
    company_id = await get_active_company_id(request, current_user)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx)")
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    imported = 0
    updated = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        try:
            nombre = str(row[0]).strip()
            rfc = str(row[1]).strip().upper() if row[1] else None
            
            vendor_data = {
                'nombre': nombre,
                'rfc': rfc,
                'email': str(row[2]).strip() if row[2] else None,
                'telefono': str(row[3]).strip() if row[3] else None,
                'direccion': str(row[4]).strip() if row[4] else None,
                'condiciones_pago': str(row[5]).strip() if len(row) > 5 and row[5] else None,
                'notas': str(row[6]).strip() if len(row) > 6 and row[6] else None
            }
            
            # Check if vendor with same RFC exists
            existing = await db.vendors.find_one({'company_id': company_id, 'rfc': rfc}, {'_id': 0}) if rfc else None
            
            if existing:
                # Update existing
                await db.vendors.update_one(
                    {'id': existing['id']},
                    {'$set': vendor_data}
                )
                updated += 1
            else:
                # Create new
                vendor = Vendor(company_id=company_id, **vendor_data)
                doc = vendor.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.vendors.insert_one(doc)
                imported += 1
                
        except Exception as e:
            errors.append(f"Fila {idx}: {str(e)}")
    
    await audit_log(company_id, 'Vendor', 'BULK', 'IMPORT', current_user['id'])
    
    return {
        'status': 'success',
        'imported': imported,
        'updated': updated,
        'errors': errors
    }

@router.post("/customers/import")
async def import_customers(request: Request, file: UploadFile = File(...), current_user: Dict = Depends(get_current_user)):
    """Import customers from Excel file"""
    from openpyxl import load_workbook
    import io
    
    company_id = await get_active_company_id(request, current_user)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos Excel (.xlsx)")
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    imported = 0
    updated = 0
    errors = []
    
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:  # Skip empty rows
            continue
        
        try:
            nombre = str(row[0]).strip()
            rfc = str(row[1]).strip().upper() if row[1] else None
            
            customer_data = {
                'nombre': nombre,
                'rfc': rfc,
                'email': str(row[2]).strip() if row[2] else None,
                'telefono': str(row[3]).strip() if row[3] else None,
                'direccion': str(row[4]).strip() if row[4] else None,
                'limite_credito': float(row[5]) if len(row) > 5 and row[5] else 0,
                'notas': str(row[6]).strip() if len(row) > 6 and row[6] else None
            }
            
            # Check if customer with same RFC exists
            existing = await db.customers.find_one({'company_id': company_id, 'rfc': rfc}, {'_id': 0}) if rfc else None
            
            if existing:
                # Update existing
                await db.customers.update_one(
                    {'id': existing['id']},
                    {'$set': customer_data}
                )
                updated += 1
            else:
                # Create new
                customer = Customer(company_id=company_id, **customer_data)
                doc = customer.model_dump()
                doc['created_at'] = doc['created_at'].isoformat()
                await db.customers.insert_one(doc)
                imported += 1
                
        except Exception as e:
            errors.append(f"Fila {idx}: {str(e)}")
    
    await audit_log(company_id, 'Customer', 'BULK', 'IMPORT', current_user['id'])
    
    return {
        'status': 'success',
        'imported': imported,
        'updated': updated,
        'errors': errors
    }

@router.post("/cfdi/{cfdi_id}/create-party")
async def create_party_from_cfdi(
    cfdi_id: str,
    party_type: str = Query(..., description="'customer' or 'vendor'"),
    nombre: str = Query(...),
    rfc: str = Query(...),
    request: Request = None,
    current_user: Dict = Depends(get_current_user)
):
    """Create a customer or vendor from CFDI data and link it"""
    company_id = await get_active_company_id(request, current_user)
    
    # Get the CFDI
    cfdi = await db.cfdis.find_one({'id': cfdi_id, 'company_id': company_id}, {'_id': 0})
    if not cfdi:
        raise HTTPException(status_code=404, detail="CFDI no encontrado")
    
    if party_type == 'customer':
        # Check if already exists
        existing = await db.customers.find_one({
            'company_id': company_id,
            'rfc': {'$regex': f'^{rfc}$', '$options': 'i'}
        }, {'_id': 0, 'id': 1})
        
        if existing:
            # Link existing
            await db.cfdis.update_one({'id': cfdi_id}, {'$set': {'customer_id': existing['id']}})
            return {'status': 'linked', 'party_id': existing['id'], 'message': 'Cliente existente vinculado'}
        
        # Create new customer
        customer = Customer(company_id=company_id, nombre=nombre, rfc=rfc.upper())
        doc = customer.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.customers.insert_one(doc)
        
        # Link to CFDI
        await db.cfdis.update_one({'id': cfdi_id}, {'$set': {'customer_id': customer.id}})
        await audit_log(company_id, 'Customer', customer.id, 'CREATE_FROM_CFDI', current_user['id'])
        
        return {
            'status': 'created',
            'party_id': customer.id,
            'party_type': 'customer',
            'nombre': nombre,
            'rfc': rfc,
            'message': f'Cliente "{nombre}" creado y vinculado'
        }
    
    elif party_type == 'vendor':
        # Check if already exists
        existing = await db.vendors.find_one({
            'company_id': company_id,
            'rfc': {'$regex': f'^{rfc}$', '$options': 'i'}
        }, {'_id': 0, 'id': 1})
        
        if existing:
            # Link existing
            await db.cfdis.update_one({'id': cfdi_id}, {'$set': {'vendor_id': existing['id']}})
            return {'status': 'linked', 'party_id': existing['id'], 'message': 'Proveedor existente vinculado'}
        
        # Create new vendor
        vendor = Vendor(company_id=company_id, nombre=nombre, rfc=rfc.upper())
        doc = vendor.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.vendors.insert_one(doc)
        
        # Link to CFDI
        await db.cfdis.update_one({'id': cfdi_id}, {'$set': {'vendor_id': vendor.id}})
        await audit_log(company_id, 'Vendor', vendor.id, 'CREATE_FROM_CFDI', current_user['id'])
        
        return {
            'status': 'created',
            'party_id': vendor.id,
            'party_type': 'vendor',
            'nombre': nombre,
            'rfc': rfc,
            'message': f'Proveedor "{nombre}" creado y vinculado'
        }
    
    else:
        raise HTTPException(status_code=400, detail="party_type debe ser 'customer' o 'vendor'")

