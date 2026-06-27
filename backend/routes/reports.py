"""Reports, dashboard and audit log routes"""
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from typing import Dict, List, Optional, Any
from datetime import datetime, timezone, timedelta
import logging

from core.database import db
from core.auth import get_current_user, get_active_company_id
from services.fx import get_fx_rate_by_date
from models.enums import UserRole
from models.audit import AuditLog

logger = logging.getLogger(__name__)

router = APIRouter()

@router.get("/audit-logs", response_model=List[AuditLog])
async def list_audit_logs(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    limit: int = Query(100, le=1000),
    skip: int = Query(0, ge=0),
    entidad: Optional[str] = Query(None, description="Filter by entity type (cfdi, payment, etc)"),
    accion: Optional[str] = Query(None, description="Filter by action (e.g. fx_corrected_to_dof)"),
    user_id: Optional[str] = Query(None, description="Filter by user id/email"),
    fecha_desde: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
    fecha_hasta: Optional[str] = Query(None, description="ISO date YYYY-MM-DD"),
):
    if current_user['role'] not in [UserRole.ADMIN, UserRole.CFO]:
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    
    company_id = await get_active_company_id(request, current_user)
    query: Dict[str, Any] = {'company_id': company_id}
    if entidad:
        query['entidad'] = entidad
    if accion:
        query['accion'] = accion
    if user_id:
        query['user_id'] = user_id
    if fecha_desde or fecha_hasta:
        ts: Dict[str, Any] = {}
        if fecha_desde:
            ts['$gte'] = fecha_desde + 'T00:00:00'
        if fecha_hasta:
            ts['$lte'] = fecha_hasta + 'T23:59:59'
        query['timestamp'] = ts
    
    logs = await db.audit_logs.find(query, {'_id': 0}).sort('timestamp', -1).skip(skip).limit(limit).to_list(limit)
    
    for log in logs:
        if isinstance(log.get('timestamp'), str):
            log['timestamp'] = datetime.fromisoformat(log['timestamp'])
    return logs


@router.get("/audit-logs/distinct")
async def audit_logs_distinct(
    request: Request,
    current_user: Dict = Depends(get_current_user),
):
    """Return distinct values of entidad/accion/user_id for the company.
    Used to populate filter dropdowns in the Bitácora UI."""
    if current_user['role'] not in [UserRole.ADMIN, UserRole.CFO]:
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
    
    company_id = await get_active_company_id(request, current_user)
    base = {'company_id': company_id}
    return {
        'entidades': sorted([v for v in await db.audit_logs.distinct('entidad', base) if v]),
        'acciones': sorted([v for v in await db.audit_logs.distinct('accion', base) if v]),
        'usuarios': sorted([v for v in await db.audit_logs.distinct('user_id', base) if v]),
    }

@router.get("/reports/dashboard")
async def get_dashboard_report(
    request: Request, 
    current_user: Dict = Depends(get_current_user),
    moneda_vista: str = Query('MXN', description='Moneda para mostrar datos'),
    bank_account_id: str = Query(None, description='Filtrar por cuenta bancaria específica'),
    fecha_desde: str = Query(None, description='Fecha inicio del rango (YYYY-MM-DD)'),
    fecha_hasta: str = Query(None, description='Fecha fin del rango (YYYY-MM-DD)')
):
    company_id = await get_active_company_id(request, current_user)
    
    # Get FX rates for conversion
    fx_rates = await db.fx_rates.find(
        {'company_id': company_id},
        {'_id': 0, 'moneda_origen': 1, 'moneda_destino': 1, 'tasa': 1}
    ).sort('fecha_vigencia', -1).to_list(100)
    
    # Build FX rates map (moneda -> MXN rate)
    fx_map = {'MXN': 1.0}
    for rate in fx_rates:
        if rate.get('moneda_destino') == 'MXN':
            fx_map[rate['moneda_origen']] = rate['tasa']
        elif rate.get('moneda_origen') == 'MXN':
            fx_map[rate['moneda_destino']] = 1 / rate['tasa']
    
    # Also check for rates stored by the forex sync service
    synced_rates = await db.fx_rates.find(
        {'company_id': company_id, 'fuente': {'$in': ['banxico', 'openexchange', 'fallback']}},
        {'_id': 0, 'moneda_origen': 1, 'tasa': 1, 'fuente': 1}
    ).sort('updated_at', -1).to_list(20)
    
    for rate in synced_rates:
        currency = rate.get('moneda_origen')
        if currency and currency not in fx_map:
            fx_map[currency] = rate['tasa']
    
    # Default rates for common currencies if not configured
    default_fx_rates = {
        'USD': 17.50,
        'EUR': 19.00,
        'GBP': 22.00,
        'JPY': 0.12,
        'CHF': 19.50,
        'CAD': 12.80,
        'CNY': 2.45
    }
    for currency, rate in default_fx_rates.items():
        if currency not in fx_map:
            fx_map[currency] = rate
    
    # Conversion factor from MXN to target currency
    target_rate = fx_map.get(moneda_vista, 1.0)
    def convert_to_target(mxn_amount):
        if moneda_vista == 'MXN':
            return mxn_amount
        return mxn_amount / target_rate
    
    # Get bank accounts
    bank_query = {'company_id': company_id}
    if bank_account_id:
        bank_query['id'] = bank_account_id
    
    bank_accounts = await db.bank_accounts.find(bank_query, {'_id': 0}).to_list(100)
    
    # Calculate saldos with conversion
    saldo_inicial_mxn = 0.0
    cash_pool_by_currency = {}
    accounts_detail = []
    
    for acc in bank_accounts:
        saldo = acc.get('saldo_inicial', 0)
        moneda = acc.get('moneda', 'MXN')
        tasa = fx_map.get(moneda, 1.0)
        saldo_mxn = saldo * tasa
        saldo_target = convert_to_target(saldo_mxn)
        
        # Cash pooling by currency
        if moneda not in cash_pool_by_currency:
            cash_pool_by_currency[moneda] = {'total': 0, 'cuentas': 0}
        cash_pool_by_currency[moneda]['total'] += saldo
        cash_pool_by_currency[moneda]['cuentas'] += 1
        
        acc['saldo_mxn'] = saldo_mxn
        acc['saldo_target'] = saldo_target
        acc['tasa_conversion'] = tasa
        
        # Risk indicators
        acc['riesgo_ocioso'] = saldo > 500000  # More than 500k idle
        acc['riesgo_bajo_saldo'] = saldo < 10000  # Less than 10k
        
        accounts_detail.append(acc)
        saldo_inicial_mxn += saldo_mxn
    
    saldo_inicial_target = convert_to_target(saldo_inicial_mxn)
    
    # Build date filter for cashflow weeks
    weeks_query = {'company_id': company_id}
    
    # Parse date filters if provided (convert YYYY-MM-DD to datetime for comparison)
    if fecha_desde:
        try:
            from datetime import datetime
            fecha_desde_dt = datetime.fromisoformat(fecha_desde + 'T00:00:00+00:00')
            weeks_query['fecha_inicio'] = {'$gte': fecha_desde_dt.isoformat()}
        except:
            pass  # Ignore invalid date format
    
    if fecha_hasta:
        try:
            from datetime import datetime
            fecha_hasta_dt = datetime.fromisoformat(fecha_hasta + 'T23:59:59+00:00')
            if 'fecha_fin' not in weeks_query:
                weeks_query['fecha_fin'] = {}
            weeks_query['fecha_fin'] = {'$lte': fecha_hasta_dt.isoformat()}
        except:
            pass  # Ignore invalid date format
    
    # Get cashflow weeks with date filter
    weeks = await db.cashflow_weeks.find(
        weeks_query,
        {'_id': 0}
    ).sort('fecha_inicio', 1).limit(52).to_list(52)  # Up to 1 year of weeks
    
    running_balance_mxn = saldo_inicial_mxn
    previous_flujo_neto = 0
    
    for idx, week in enumerate(weeks):
        txn_query = {'company_id': company_id, 'cashflow_week_id': week['id']}
        if bank_account_id:
            txn_query['bank_account_id'] = bank_account_id
        
        transactions = await db.transactions.find(txn_query, {'_id': 0}).to_list(1000)
        
        total_ingresos_mxn = sum(t['monto'] for t in transactions if t['tipo_transaccion'] == 'ingreso')
        total_egresos_mxn = sum(t['monto'] for t in transactions if t['tipo_transaccion'] == 'egreso')
        flujo_neto_mxn = total_ingresos_mxn - total_egresos_mxn
        
        week['saldo_inicial_mxn'] = running_balance_mxn
        week['saldo_inicial'] = convert_to_target(running_balance_mxn)
        week['total_ingresos_mxn'] = total_ingresos_mxn
        week['total_ingresos'] = convert_to_target(total_ingresos_mxn)
        week['total_egresos_mxn'] = total_egresos_mxn
        week['total_egresos'] = convert_to_target(total_egresos_mxn)
        week['flujo_neto_mxn'] = flujo_neto_mxn
        week['flujo_neto'] = convert_to_target(flujo_neto_mxn)
        week['saldo_final_mxn'] = running_balance_mxn + flujo_neto_mxn
        week['saldo_final'] = convert_to_target(week['saldo_final_mxn'])
        
        # Variance calculation (vs previous week)
        if idx > 0:
            week['varianza_flujo'] = flujo_neto_mxn - previous_flujo_neto
            week['varianza_pct'] = (week['varianza_flujo'] / abs(previous_flujo_neto) * 100) if previous_flujo_neto != 0 else 0
        else:
            week['varianza_flujo'] = 0
            week['varianza_pct'] = 0
        
        previous_flujo_neto = flujo_neto_mxn
        running_balance_mxn = week['saldo_final_mxn']
    
    # Calculate trends and risk indicators
    if len(weeks) >= 4:
        recent_flows = [w['flujo_neto_mxn'] for w in weeks[-4:]]
        trend_direction = 'up' if recent_flows[-1] > recent_flows[0] else 'down' if recent_flows[-1] < recent_flows[0] else 'stable'
        avg_flow = sum(recent_flows) / len(recent_flows)
    else:
        trend_direction = 'stable'
        avg_flow = 0
    
    # Risk indicators
    saldo_final_proyectado = weeks[-1]['saldo_final_mxn'] if weeks else saldo_inicial_mxn
    risk_indicators = {
        'liquidez_critica': saldo_final_proyectado < 50000,
        'tendencia_negativa': trend_direction == 'down' and avg_flow < 0,
        'saldos_ociosos': sum(1 for acc in accounts_detail if acc.get('riesgo_ocioso', False)),
        'cuentas_bajo_saldo': sum(1 for acc in accounts_detail if acc.get('riesgo_bajo_saldo', False)),
        'semanas_con_deficit': sum(1 for w in weeks if w.get('flujo_neto_mxn', 0) < 0)
    }
    
    # KPIs
    total_transactions = await db.transactions.count_documents({'company_id': company_id})
    total_cfdis = await db.cfdis.count_documents({'company_id': company_id})
    total_reconciliations = await db.reconciliations.count_documents({'company_id': company_id})
    total_customers = await db.customers.count_documents({'company_id': company_id})
    total_vendors = await db.vendors.count_documents({'company_id': company_id})
    
    return {
        'moneda_vista': moneda_vista,
        'cashflow_weeks': weeks,
        'saldo_inicial_bancos': saldo_inicial_target,
        'saldo_inicial_bancos_mxn': saldo_inicial_mxn,
        'saldo_final_proyectado': convert_to_target(saldo_final_proyectado),
        'saldo_final_proyectado_mxn': saldo_final_proyectado,
        'bank_accounts': accounts_detail,
        'fx_rates_used': fx_map,
        'cash_pool': cash_pool_by_currency,
        'trend': {
            'direction': trend_direction,
            'avg_flow_4w': convert_to_target(avg_flow),
            'avg_flow_4w_mxn': avg_flow
        },
        'risk_indicators': risk_indicators,
        'kpis': {
            'total_transactions': total_transactions,
            'total_cfdis': total_cfdis,
            'total_reconciliations': total_reconciliations,
            'total_customers': total_customers,
            'total_vendors': total_vendors
        }
    }


@router.get("/reports/dashboard-from-payments")
async def get_dashboard_from_payments(
    request: Request,
    current_user: Dict = Depends(get_current_user),
    moneda_vista: str = Query('MXN', description='Moneda para mostrar datos'),
    bank_account_id: Optional[str] = Query(None, description='Filtrar por cuenta bancaria específica'),
    fecha_inicio: Optional[str] = Query(None, description='Inicio del rango (YYYY-MM-DD) — genera semanas que cubran el rango'),
    fecha_fin: Optional[str] = Query(None, description='Fin del rango (YYYY-MM-DD)')
):
    """
    Dashboard alternativo que genera datos directamente desde pagos reales.
    Usa la misma lógica que CashflowProjections para consistencia.
    Sin fecha_inicio/fecha_fin genera la ventana default de 13 semanas
    (4 pasadas + actual + 8 futuras); con rango genera las semanas que lo cubran.
    """
    from datetime import datetime, timedelta
    
    company_id = await get_active_company_id(request, current_user)
    
    # Get FX rates
    fx_rates = await db.fx_rates.find({'company_id': company_id}, {'_id': 0}).to_list(100)
    fx_map = {'MXN': 1.0, 'USD': 17.5, 'EUR': 20.0}
    for rate in fx_rates:
        if rate.get('moneda_destino') == 'MXN' and rate.get('tasa'):
            fx_map[rate['moneda_origen']] = rate['tasa']
    
    def convert_to_mxn(amount, currency):
        return amount * fx_map.get(currency, 1)
    
    def convert_from_mxn(amount_mxn, target_currency):
        """Convert MXN to target currency"""
        if target_currency == 'MXN':
            return amount_mxn
        rate = fx_map.get(target_currency, 1)
        return amount_mxn / rate if rate else amount_mxn
    
    def to_display_currency(amount_mxn):
        """Convert MXN amount to display currency"""
        return convert_from_mxn(amount_mxn, moneda_vista)
    
    # Get bank account balances - only active accounts, same filter as bank-accounts/summary
    accounts_query = {'company_id': company_id, 'activo': True}
    if bank_account_id:
        accounts_query['id'] = bank_account_id
    
    accounts = await db.bank_accounts.find(accounts_query, {'_id': 0}).to_list(50)
    
    # Calculate initial balance using fecha_saldo for FX rate
    saldo_bancos_mxn = 0
    selected_account_moneda = 'MXN'
    selected_account_saldo = 0
    
    for acc in accounts:
        saldo = acc.get('saldo_inicial', 0) or 0
        moneda = acc.get('moneda', 'MXN')
        fecha_saldo = acc.get('fecha_saldo')
        
        # Use the FX rate from fecha_saldo date
        if fecha_saldo and moneda != 'MXN':
            if isinstance(fecha_saldo, str):
                fecha_saldo = datetime.fromisoformat(fecha_saldo.replace('Z', '+00:00').split('+')[0])
            rate = await get_fx_rate_by_date(company_id, moneda, fecha_saldo)
        else:
            rate = fx_map.get(moneda, 1) if moneda != 'MXN' else 1
        
        saldo_mxn = saldo * rate if moneda != 'MXN' else saldo
        saldo_bancos_mxn += saldo_mxn
        
        if bank_account_id and acc.get('id') == bank_account_id:
            selected_account_moneda = moneda
            selected_account_saldo = saldo
    
    # Get all bank transactions
    bank_txns_query = {'company_id': company_id}
    if bank_account_id:
        bank_txns_query['bank_account_id'] = bank_account_id
    
    bank_txns = await db.bank_transactions.find(bank_txns_query, {'_id': 0, 'id': 1, 'conciliado': 1, 'bank_account_id': 1}).to_list(20000)
    reconciled_ids = set(t['id'] for t in bank_txns if t.get('conciliado') == True)
    bank_txn_to_account = {t['id']: t.get('bank_account_id') for t in bank_txns}

    all_payments = await db.payments.find({'company_id': company_id, 'estatus': 'completado'}, {'_id': 0}).to_list(10000)

    # Si la empresa usa Alegra, incluir bank_transactions como pagos
    company_doc = await db.companies.find_one({'id': company_id}, {'_id': 0, 'alegra_connected': 1})
    if company_doc and company_doc.get('alegra_connected'):
        bank_txns_alegra = await db.bank_transactions.find(
            {'company_id': company_id, 'source': 'alegra', 'es_real': True},
            {'_id': 0, 'id': 1, 'tipo': 1, 'monto': 1, 'moneda': 1,
             'moneda_original': 1, 'monto_original': 1, 'fecha': 1,
             'contacto': 1, 'descripcion': 1, 'cuenta_bancaria': 1}
        ).to_list(10000)

        for bt in bank_txns_alegra:
            all_payments.append({
                'company_id': company_id,
                'tipo': 'cobro' if bt.get('tipo') == 'deposito' else 'pago',
                'monto': bt.get('monto', 0),
                'moneda': bt.get('moneda_original') or bt.get('moneda', 'MXN'),
                'fecha_pago': bt.get('fecha'),
                'estatus': 'completado',
                'bank_transaction_id': bt.get('id'),
                'beneficiario': bt.get('contacto') or bt.get('descripcion', ''),
                '_from_alegra': True,
            })

    logger.info(f"[dashboard-debug] payments de db.payments: {len(all_payments)}")
    logger.info(f"[dashboard-debug] company alegra_connected: {company_doc}")

    # Filter to valid payments
    # Para empresas que usan Alegra con bank_transactions como fuente principal,
    # incluir TODOS los pagos completados independientemente de si están conciliados
    uses_alegra = company_doc and company_doc.get('alegra_connected')
    
    if uses_alegra:
        # Alegra: incluir todos los pagos completados
        payments = [p for p in all_payments if p.get('estatus') == 'completado' or p.get('_from_alegra')]
    else:
        # Contalink/manual: solo incluir pagos conciliados o sin bank_transaction
        payments = [p for p in all_payments
                    if not p.get('bank_transaction_id')
                    or p.get('bank_transaction_id') in reconciled_ids
                    or p.get('_from_alegra') == True]

    logger.info(f"[dashboard-debug] payments después de filtro: {len(payments)}")
    logger.info(f"[dashboard-debug] sample payment: {payments[0] if payments else 'VACÍO'}")

    # If filtering by bank account, only include payments for that account
    if bank_account_id:
        payments = [p for p in payments if p.get('bank_transaction_id') and bank_txn_to_account.get(p['bank_transaction_id']) == bank_account_id]
    
    # Get categories for USD operations
    categories = await db.categories.find({'company_id': company_id}, {'_id': 0}).to_list(100)
    compra_usd_id = next((c['id'] for c in categories if 'compra' in c.get('nombre', '').lower() and 'usd' in c.get('nombre', '').lower()), None)
    venta_usd_id = next((c['id'] for c in categories if 'venta' in c.get('nombre', '').lower() and 'usd' in c.get('nombre', '').lower()), None)
    
    # Generate weeks window
    today = datetime.now()
    # Find Monday of current week
    days_since_monday = today.weekday()
    current_monday = today - timedelta(days=days_since_monday)

    def _parse_fecha(s):
        try:
            return datetime.fromisoformat(s) if s else None
        except ValueError:
            return None

    rango_inicio = _parse_fecha(fecha_inicio)
    rango_fin    = _parse_fecha(fecha_fin)

    if rango_inicio and rango_fin and rango_fin >= rango_inicio:
        # Ventana dinámica: semanas (lunes a lunes) que cubren el rango pedido (cap 60)
        start_monday = rango_inicio - timedelta(days=rango_inicio.weekday())
        num_weeks = min(int((rango_fin - start_monday).days // 7) + 1, 60)
    else:
        # Default: 13 semanas (4 pasadas, actual, 8 futuras)
        start_monday = current_monday - timedelta(weeks=4)
        num_weeks = 13

    # ── Saldo base: usa ancla histórica más cercana a fecha_inicio ────────────
    # Si hay bank_account_history para la fecha del rango, usa ese saldo verificado.
    # Si no, usa el saldo_inicial actual de bank_accounts (el más reciente).
    # Esto garantiza que filtrar por enero 2026 use el saldo de enero, no el de mayo.
    all_active_accs = await db.bank_accounts.find(
        {'company_id': company_id, 'activo': True}, {'_id': 0}
    ).to_list(1000)

    # Fecha de referencia: inicio del rango pedido (o hoy si no hay rango)
    fecha_ref = (rango_inicio or datetime.now()).strftime('%Y-%m-%d')

    # Para cada cuenta, buscar el saldo histórico más cercano ANTERIOR o IGUAL a fecha_ref
    bank_base_mxn = 0.0
    for acc in all_active_accs:
        acct_id = acc['id']
        moneda_acc = acc.get('moneda', 'MXN')
        tc = fx_map.get(moneda_acc, 1.0)

        # Buscar ancla histórica <= fecha_ref, la más reciente
        hist = await db.bank_account_history.find_one(
            {'account_id': acct_id, 'company_id': company_id, 'fecha': {'$lte': fecha_ref}},
            {'_id': 0, 'saldo': 1, 'fecha': 1},
            sort=[('fecha', -1)],
        )
        if hist:
            saldo_ref = float(hist.get('saldo', 0) or 0)
            logger.info(f"[dashboard-anchor] cuenta {acct_id[:8]} {moneda_acc}: "
                        f"ancla {hist['fecha']} → ${saldo_ref:,.2f} (ref={fecha_ref})")
        else:
            # Sin historial: usar saldo_inicial actual
            saldo_ref = float(acc.get('saldo_inicial', 0) or 0)
            logger.info(f"[dashboard-anchor] cuenta {acct_id[:8]} {moneda_acc}: "
                        f"sin historial, usando saldo_inicial ${saldo_ref:,.2f}")

        bank_base_mxn += saldo_ref if moneda_acc == 'MXN' else saldo_ref * tc

    saldo_actual_mxn = bank_base_mxn
    for p in all_payments:
        fecha_str = p.get('fecha_pago') or p.get('fecha_vencimiento')
        if not fecha_str:
            continue
        try:
            fecha_dt = datetime.fromisoformat(fecha_str.replace('Z', '+00:00').split('+')[0])
        except (ValueError, AttributeError):
            continue
        # Excluir traspasos entre cuentas del cálculo de saldo
        concepto_p = (p.get('concepto') or p.get('descripcion') or '').lower()
        TRASPASO_KW = ['operacion cambios', 'operación cambios', 'cambio de divisa',
                       'traspaso', 'retiro por operacion', 'deposito por operacion']
        if any(kw in concepto_p for kw in TRASPASO_KW):
            continue
        monto_mxn = convert_to_mxn(p.get('monto', 0), p.get('moneda', 'MXN'))
        cat_id = p.get('category_id')
        # Mismo fix: solo enrutar a venta/compra_usd si el id existe (None != None guard)
        if venta_usd_id and cat_id == venta_usd_id:
            saldo_actual_mxn += monto_mxn
        elif compra_usd_id and cat_id == compra_usd_id:
            saldo_actual_mxn -= monto_mxn
        elif p.get('tipo') in ('cobro', 'ingreso'):
            saldo_actual_mxn += monto_mxn
        else:
            saldo_actual_mxn -= monto_mxn

    for _acc in all_active_accs:
        print(
            f"[cuenta_debug] nombre={_acc.get('nombre')} banco={_acc.get('banco')} "
            f"saldo_inicial={_acc.get('saldo_inicial')} moneda={_acc.get('moneda')} "
            f"fecha_saldo={_acc.get('fecha_saldo')} activo={_acc.get('activo')}",
            flush=True
        )
    print(
        f"[saldo_actual_debug] company={company_id} "
        f"cuentas_activas={len(all_active_accs)} bank_base_mxn={bank_base_mxn:,.0f} "
        f"all_payments={len(all_payments)} saldo_actual_mxn={saldo_actual_mxn:,.0f} "
        f"saldo_bancos_mxn={saldo_bancos_mxn:,.0f}",
        flush=True
    )

    # fecha_saldo más antigua entre TODAS las cuentas activas (para aviso de desactualización)
    fechas_saldo = []
    for acc in all_active_accs:
        fs = acc.get('fecha_saldo')
        if fs:
            try:
                if isinstance(fs, str):
                    fs = datetime.fromisoformat(fs.replace('Z', '+00:00').split('+')[0])
                fechas_saldo.append(fs)
            except (ValueError, AttributeError):
                pass
    fecha_saldo_bancos_iso = min(fechas_saldo).isoformat() if fechas_saldo else None

    # ── Proyecciones CxC/CxP por semana ──────────────────────────────────────
    # La colección cxc_proyecciones usa numeración absoluta desde S1 = primer
    # lunes del año en curso. Calculamos model_start para convertir cada
    # week_start del loop a su semana_label absoluta y así cruzar los montos.
    proy_raw = await db.cxc_proyecciones.find(
        {"company_id": company_id, "semana": {"$ne": None}}, {"_id": 0}
    ).to_list(2000)

    _year_start = datetime(today.year, 1, 1)
    _model_start = _year_start - timedelta(days=_year_start.weekday())  # lunes previo al 1-ene

    proy_por_semana: dict = {}
    for _p in proy_raw:
        _semana = _p.get("semana")
        _tipo   = _p.get("tipo", "cxc")
        _monto  = convert_to_mxn(float(_p.get("monto", 0) or 0), _p.get("moneda", "MXN"))
        if _semana not in proy_por_semana:
            proy_por_semana[_semana] = {"cxc": 0.0, "cxp": 0.0}
        proy_por_semana[_semana][_tipo] = proy_por_semana[_semana][_tipo] + _monto

    weeks_data = []
    running_balance = saldo_bancos_mxn  # base del gráfico (saldo apertura + pagos conciliados)

    # ── Saldo inicial del modelo ──────────────────────────────────────────────
    # El running_balance debe partir del saldo bancario real AJUSTADO por los
    # movimientos ocurridos ANTES del inicio de la ventana del modelo.
    # Así si el saldo bancario es de mayo-30 y el modelo empieza en junio,
    # el saldo inicial del modelo refleja la realidad.
    #
    # Lógica:
    # 1. Tomar el saldo bancario registrado (fecha_saldo en BD)
    # 2. Sumar/restar los pagos ocurridos ENTRE fecha_saldo y start_monday
    # 3. El resultado es el saldo real al inicio de la ventana del modelo

    # Determinar la fecha_saldo más reciente de las cuentas
    fecha_saldo_base = None
    for acc in all_active_accs:
        fs_raw = acc.get('fecha_saldo')
        if fs_raw:
            try:
                fs = datetime.fromisoformat(str(fs_raw).replace('Z', '+00:00').split('+')[0])
                if fecha_saldo_base is None or fs > fecha_saldo_base:
                    fecha_saldo_base = fs
            except (ValueError, AttributeError):
                pass

    # Ajustar el saldo bancario con pagos entre fecha_saldo y start_monday
    adjusted_balance = saldo_bancos_mxn
    if fecha_saldo_base and fecha_saldo_base < start_monday:
        for p in payments:
            fecha_str = p.get('fecha_pago') or p.get('fecha_vencimiento')
            if not fecha_str:
                continue
            try:
                fecha_dt = datetime.fromisoformat(fecha_str.replace('Z', '+00:00').split('+')[0])
            except (ValueError, AttributeError):
                continue
            # Solo pagos entre fecha_saldo y start_monday
            if not (fecha_saldo_base <= fecha_dt < start_monday):
                continue
            concepto_adj = (p.get('concepto') or p.get('descripcion') or '').lower()
            TRASPASO_KW2 = ['operacion cambios', 'operación cambios', 'cambio de divisa',
                            'traspaso', 'retiro por operacion', 'deposito por operacion']
            if any(kw in concepto_adj for kw in TRASPASO_KW2):
                continue
            monto_adj = convert_to_mxn(p.get('monto', 0), p.get('moneda', 'MXN'))
            cat_id_adj = p.get('category_id')
            if venta_usd_id and cat_id_adj == venta_usd_id:
                adjusted_balance += monto_adj
            elif compra_usd_id and cat_id_adj == compra_usd_id:
                adjusted_balance -= monto_adj
            elif p.get('tipo') in ('cobro', 'ingreso'):
                adjusted_balance += monto_adj
            else:
                adjusted_balance -= monto_adj

    running_balance = adjusted_balance
    logger.info(f"[dashboard] saldo_base={saldo_bancos_mxn:,.0f} fecha_saldo={fecha_saldo_base} "
                f"start_monday={start_monday} adjusted_balance={adjusted_balance:,.0f}")

    for i in range(num_weeks):
        week_start = start_monday + timedelta(weeks=i)
        week_end = week_start + timedelta(days=7)
        is_past = week_end <= today
        is_current = week_start <= today < week_end

        # Semana absoluta del año (S1 = primer lunes del año) para lookup en cxc_proyecciones
        abs_semana_label = f"S{(week_start - _model_start).days // 7 + 1}"

        # Filter payments for this week
        week_payments = [p for p in payments if p.get('fecha_pago')]
        week_payments = [p for p in week_payments if week_start <= datetime.fromisoformat(p['fecha_pago'].replace('Z', '+00:00').split('+')[0]) < week_end]

        if i == 0:  # solo para la primera semana
            logger.info(f"[dash-fecha] week_start={week_start} week_end={week_end} total_payments={len(payments)} week_payments={len(week_payments)}")
            if payments:
                p = payments[0]
                logger.info(f"[dash-fecha] sample fecha_pago='{p.get('fecha_pago')}' tipo='{p.get('tipo')}'")

        # Calculate totals excluding USD operations
        ingresos = 0
        egresos = 0
        venta_usd = 0
        compra_usd = 0

        # Palabras clave de traspasos entre cuentas — no son egresos/ingresos reales
        TRASPASO_KEYWORDS = [
            'operacion cambios', 'operación cambios', 'cambio de divisa',
            'traspaso', 'transferencia entre cuentas', 'compra de dolares',
            'compra de dólares', 'venta de dolares', 'venta de dólares',
            'retiro por operacion', 'deposito por operacion',
        ]

        for p in week_payments:
            # Excluir traspasos entre cuentas (USD↔MXN) — se netean solos
            concepto = (p.get('concepto') or p.get('descripcion') or '').lower()
            if any(kw in concepto for kw in TRASPASO_KEYWORDS):
                continue

            moneda_pago = p.get('moneda', 'MXN')
            monto_raw = p.get('monto', 0)
            if p.get('monto_mxn'):
                monto_mxn = p.get('monto_mxn')
            elif moneda_pago == 'MXN':
                monto_mxn = monto_raw
            else:
                monto_mxn = convert_to_mxn(monto_raw, moneda_pago)
            cat_id = p.get('category_id')

            # IMPORTANTE: solo enrutar a venta_usd/compra_usd si el id existe
            # (None == None es True en Python)
            if venta_usd_id and cat_id == venta_usd_id:
                venta_usd += monto_mxn
            elif compra_usd_id and cat_id == compra_usd_id:
                compra_usd += monto_mxn
            elif p.get('tipo') in ('cobro', 'ingreso'):
                ingresos += monto_mxn
            else:
                egresos += monto_mxn

        # Sumar proyecciones CxC (cobros) y CxP (pagos) para semanas no pasadas
        proy_semana = proy_por_semana.get(abs_semana_label, {})
        ingreso_cxc = proy_semana.get("cxc", 0.0) if not is_past else 0.0
        egreso_cxp  = proy_semana.get("cxp", 0.0) if not is_past else 0.0

        flujo_neto = ingresos - egresos + venta_usd - compra_usd + ingreso_cxc - egreso_cxp
        saldo_final = running_balance + flujo_neto

        weeks_data.append({
            'week_num': i + 1,
            'week_label': f"S{i + 1}",
            'abs_semana': abs_semana_label,
            'date_label': week_start.strftime('%d %b'),
            'fecha_inicio': week_start.isoformat(),
            'fecha_fin': week_end.isoformat(),
            'is_past': is_past,
            'is_current': is_current,
            'ingresos': round(ingresos, 2),
            'egresos': round(egresos, 2),
            'ingreso_cxc': round(ingreso_cxc, 2),
            'egreso_cxp': round(egreso_cxp, 2),
            'venta_usd': round(venta_usd, 2),
            'compra_usd': round(compra_usd, 2),
            'flujo_neto': round(flujo_neto, 2),
            'saldo_inicial': round(running_balance, 2),
            'saldo_final': round(saldo_final, 2),
            'num_payments': len(week_payments)
        })

        running_balance = saldo_final

    # [DEBUG-TEMP] Auditar cascadeo de saldo por semana con proyecciones CxC/CxP
    for _w in weeks_data:
        logger.info(
            f"[saldo_debug] {_w['week_label']} ({_w['abs_semana']} / {_w['date_label']}) "
            f"is_past={_w['is_past']} is_current={_w['is_current']} "
            f"ingresos={_w['ingresos']:,.0f} egresos={_w['egresos']:,.0f} "
            f"cxc={_w['ingreso_cxc']:,.0f} cxp={_w['egreso_cxp']:,.0f} "
            f"flujo={_w['flujo_neto']:,.0f} saldo_final={_w['saldo_final']:,.0f}"
        )
    logger.info(
        f"[saldo_debug] saldo_proyectado final={running_balance:,.0f} | "
        f"proyecciones cargadas: {len(proy_por_semana)} semanas "
        f"({sum(v['cxc'] for v in proy_por_semana.values()):,.0f} CxC / "
        f"{sum(v['cxp'] for v in proy_por_semana.values()):,.0f} CxP)"
    )

    # Calculate varianza (change vs previous week) for each week
    for i, week in enumerate(weeks_data):
        if i == 0:
            week['varianza'] = 0
            week['varianza_pct'] = 0
        else:
            prev_week = weeks_data[i - 1]
            week['varianza'] = round(week['flujo_neto'] - prev_week['flujo_neto'], 2)
            if prev_week['flujo_neto'] != 0:
                week['varianza_pct'] = round((week['varianza'] / abs(prev_week['flujo_neto'])) * 100, 1)
            else:
                week['varianza_pct'] = 0
    
    # Calculate KPIs
    past_weeks = [w for w in weeks_data if w['is_past'] or w['is_current']]
    total_ingresos = sum(w['ingresos'] + w['venta_usd'] for w in past_weeks)
    total_egresos  = sum(w['egresos']  + w['compra_usd'] for w in past_weeks)

    # Burn rate: usar solo semanas con actividad real (egresos > 0)
    # para evitar que semanas vacías diluyan el promedio
    active_weeks = [w for w in past_weeks if (w['egresos'] + w['compra_usd']) > 0]
    burn_rate = total_egresos / len(active_weeks) if active_weeks else 0

    # Runway: saldo actual ÷ burn rate semanal
    runway_weeks = saldo_actual_mxn / burn_rate if burn_rate > 0 else float('inf')
    
    # Find critical week (first week with negative balance)
    critical_week = None
    for w in weeks_data:
        if w['saldo_final'] < 0:
            critical_week = w['week_label']
            break
    
    # Build cash pool by currency
    cash_pool = {}
    for acc in accounts:
        moneda = acc.get('moneda', 'MXN')
        saldo = acc.get('saldo_inicial', 0) or 0
        if moneda not in cash_pool:
            cash_pool[moneda] = {'total': 0, 'cuentas': 0}
        cash_pool[moneda]['total'] += saldo
        cash_pool[moneda]['cuentas'] += 1
    
    # Calculate movements per account for current month
    # Group payments by bank account to calculate saldo_final
    account_movements = {}
    for p in payments:
        bank_txn_id = p.get('bank_transaction_id')
        if bank_txn_id:
            acc_id = bank_txn_to_account.get(bank_txn_id)
            if acc_id:
                if acc_id not in account_movements:
                    account_movements[acc_id] = {'ingresos': 0, 'egresos': 0, 'count': 0}
                if p.get('tipo') == 'cobro':
                    account_movements[acc_id]['ingresos'] += p.get('monto', 0)
                else:
                    account_movements[acc_id]['egresos'] += p.get('monto', 0)
                account_movements[acc_id]['count'] += 1
    
    # Build bank accounts detail with calculated saldo_final
    bank_accounts_detail = []
    for acc in accounts:
        saldo_inicial = acc.get('saldo_inicial', 0) or 0
        moneda = acc.get('moneda', 'MXN')
        acc_id = acc.get('id')
        
        # Get movements for this account
        movements = account_movements.get(acc_id, {'ingresos': 0, 'egresos': 0, 'count': 0})
        
        # Calculate saldo_final = saldo_inicial + ingresos - egresos
        saldo_final = saldo_inicial + movements['ingresos'] - movements['egresos']
        
        saldo_inicial_mxn = convert_to_mxn(saldo_inicial, moneda)
        saldo_final_mxn = convert_to_mxn(saldo_final, moneda)
        
        bank_accounts_detail.append({
            'id': acc_id,
            'nombre': acc.get('nombre'),
            'banco': acc.get('banco'),
            'numero_cuenta': acc.get('numero_cuenta'),
            'moneda': moneda,
            'saldo_inicial': saldo_inicial,
            'saldo_final': round(saldo_final, 2),
            'saldo_inicial_mxn': saldo_inicial_mxn,
            'saldo_final_mxn': saldo_final_mxn,
            'saldo_display': round(to_display_currency(saldo_final_mxn), 2),
            'ingresos': round(movements['ingresos'], 2),
            'egresos': round(movements['egresos'], 2),
            'num_movimientos': movements['count'],
            'riesgo': 'bajo' if saldo_final_mxn > 50000 else 'medio' if saldo_final_mxn > 10000 else 'alto'
        })
    
    # Convert weeks data to display currency
    for week in weeks_data:
        week['ingresos_display'] = round(to_display_currency(week['ingresos']), 2)
        week['egresos_display'] = round(to_display_currency(week['egresos']), 2)
        week['flujo_neto_display'] = round(to_display_currency(week['flujo_neto']), 2)
        week['saldo_inicial_display'] = round(to_display_currency(week['saldo_inicial']), 2)
        week['saldo_final_display'] = round(to_display_currency(week['saldo_final']), 2)
        week['varianza_display'] = round(to_display_currency(week.get('varianza', 0)), 2)
    
    # Get FX rate for display
    fx_rate_display = fx_map.get(moneda_vista, 1) if moneda_vista != 'MXN' else 1
    
    # Build response with account filter info
    response = {
        'moneda_vista': moneda_vista,
        'fx_rate': fx_rate_display,
        'saldo_bancos': round(to_display_currency(saldo_bancos_mxn), 2),         # apertura (estático)
        'saldo_actual': round(to_display_currency(saldo_actual_mxn), 2),          # calculado (apertura + flujos pre-ventana)
        'fecha_saldo_bancos': fecha_saldo_bancos_iso,                              # para aviso de desactualización
        'saldo_proyectado': round(to_display_currency(running_balance), 2),
        'total_ingresos': round(to_display_currency(total_ingresos), 2),
        'total_egresos': round(to_display_currency(total_egresos), 2),
        'burn_rate': round(to_display_currency(burn_rate), 2),
        'runway_weeks': round(runway_weeks, 1) if runway_weeks != float('inf') else None,
        'critical_week': critical_week,
        'cobranza_vs_pagos': round((total_ingresos / total_egresos * 100), 1) if total_egresos > 0 else 100,
        'weeks': weeks_data,
        'cash_pool': cash_pool,
        'bank_accounts': bank_accounts_detail,
        'kpis': {
            'total_payments': len(payments),
            'total_cfdis': await db.cfdis.count_documents({'company_id': company_id}),
            'total_customers': await db.customers.count_documents({'company_id': company_id}),
            'total_vendors': await db.vendors.count_documents({'company_id': company_id})
        }
    }
    
    # Add filtered account info if filtering by specific account
    if bank_account_id and len(accounts) > 0:
        acc = accounts[0]
        response['filtered_account'] = {
            'id': acc.get('id'),
            'nombre': acc.get('nombre'),
            'banco': acc.get('banco'),
            'moneda': acc.get('moneda'),
            'saldo_inicial': acc.get('saldo_inicial', 0) or 0,
            'saldo_inicial_display': round(to_display_currency(convert_to_mxn(acc.get('saldo_inicial', 0) or 0, acc.get('moneda', 'MXN'))), 2),
            'num_movements': len(payments)
        }
    
    return response


# ================== ENDPOINTS AVANZADOS - FASE 2 ==================

# Importar servicios avanzados
from advanced_services import PredictiveAnalysisService, AutoReconciliationService, AlertService
from integration_services import SATScraperService, BankAPIService, SATCredentialManager

# ===== AN\u00c1LISIS PREDICTIVO CON IA =====


# ===== PDF MEJORADO - Reporte Ejecutivo con Gráficas =====
from fastapi import Body
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import io

class PDFMejoradoRequest(BaseModel):
    empresa: str = "Mi Empresa"
    rfc: str = ""
    periodo: str = ""
    ingresos: float = 0
    costo_ventas: float = 0
    utilidad_bruta: float = 0
    gastos_op: float = 0
    ebitda: float = 0
    utilidad_neta: float = 0
    activo_total: float = 0
    activo_circ: float = 0
    activo_fijo: float = 0
    pasivo_total: float = 0
    pasivo_circ: float = 0
    pasivo_lp: float = 0
    capital: float = 0
    margen_bruto: float = 0
    margen_ebitda: float = 0
    margen_op: float = 0
    margen_neto: float = 0
    roic: float = 0
    roe: float = 0
    roa: float = 0
    roce: float = 0
    razon_circ: float = 0
    prueba_acida: float = 0
    razon_ef: float = 0
    capital_trabajo: float = 0
    cash_runway: float = 0
    dso: float = 0
    dpo: float = 0
    dio: float = 0
    ccc: float = 0
    deuda_capital: float = 0
    deuda_activos: float = 0
    deuda_ebitda: float = 0
    cobertura: float = 0
    apalancamiento: float = 0
    top_cxc: list = []
    top_cxp: list = []
    ai_analysis: dict = {}
    trends: list = []
    report_type: str = 'ejecutivo'


@router.post("/reports/pdf-mejorado")
async def generate_pdf_mejorado(
    request: Request,
    data: PDFMejoradoRequest,
    current_user: Dict = Depends(get_current_user),
):
    """
    Genera el Reporte Ejecutivo o CFO con gráficas y análisis profundo.
    Usa report_type='ejecutivo' (default) o 'cfo'. Devuelve PDF binario.
    """
    try:
        if data.report_type == 'cfo':
            from services.pdf_generator import build_pdf_cfo
            pdf_buffer = build_pdf_cfo(data.dict())
            empresa_safe = data.empresa.replace(" ", "_").replace("/", "-")
            filename = f"Reporte_CFO_{empresa_safe}_{data.periodo}.pdf"
        else:
            from services.pdf_generator import build_pdf_mejorado
            pdf_buffer = build_pdf_mejorado(data.dict())
            empresa_safe = data.empresa.replace(" ", "_").replace("/", "-")
            filename = f"Reporte_Ejecutivo_{empresa_safe}_{data.periodo}.pdf"

        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )

    except Exception as e:
        logger.error(f"Error generando PDF Mejorado: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generando PDF: {str(e)}")


# ═══════════════════════════════════════════════════════════════
# EXCEL CORPORATIVO — Reporte financiero completo en 5 hojas
# ═══════════════════════════════════════════════════════════════
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import io

class ExcelCorporativoRequest(BaseModel):
    empresa: str
    rfc: str = ""
    periodo: str
    income_statement: dict = {}
    balance_sheet: dict = {}
    metrics: dict = {}
    ai_analysis: dict = {}
    trends: list = []

@router.post("/excel-corporativo")
async def generate_excel_corporativo(
    data: ExcelCorporativoRequest,
    request: Request,
    current_user: Dict = Depends(get_current_user)
):
    """Genera reporte financiero ejecutivo en Excel con 5 hojas profesionales."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import date

        wb = openpyxl.Workbook()

        # ── Colores ────────────────────────────────────────────
        DARK_BLUE  = "1A3A5C"
        MID_BLUE   = "2E6DA4"
        LIGHT_BLUE = "D6E8F7"
        GREEN      = "217346"
        LIGHT_GREEN= "E2EFDA"
        RED        = "C00000"
        LIGHT_RED  = "FFE0E0"
        ORANGE     = "ED7D31"
        GRAY_H     = "F2F2F2"
        WHITE      = "FFFFFF"

        def hdr(ws, row, col, val, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, h_align="center", span=None):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(bold=bold, color=fg, size=size, name="Arial")
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal=h_align, vertical="center", wrap_text=True)
            if span:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
            return cell

        def lbl(ws, row, col, v, bold=False, fg="000000", bg=None, span=None, italic=False):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(bold=bold, color=fg, size=10, name="Arial", italic=italic)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            if span: ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
            return cell

        def num(ws, row, col, v, fmt='$#,##0', bold=False, fg="000000", bg=None):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = Font(bold=bold, color=fg, size=10, name="Arial")
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if bg: cell.fill = PatternFill("solid", fgColor=bg)
            return cell

        # Extraer datos del income_statement
        inc = data.income_statement
        bal = data.balance_sheet
        mtr = data.metrics
        empresa = data.empresa
        periodo = data.periodo
        rfc     = data.rfc
        hoy     = date.today().strftime("%d/%m/%Y")

        # Valores clave
        ingresos       = inc.get("ingresos_netos", inc.get("ingresos", 0)) or 0
        util_bruta     = inc.get("utilidad_bruta", 0) or 0
        ebitda         = inc.get("ebitda", inc.get("utilidad_operativa", 0)) or 0
        util_neta      = inc.get("utilidad_neta", inc.get("resultado_neto", 0)) or 0
        costo_venta    = inc.get("costo_ventas", inc.get("costo_venta", 0)) or 0
        gastos_op      = inc.get("gastos_operativos", inc.get("gastos_operacion", 0)) or 0
        gastos_fin     = inc.get("gastos_financieros", 0) or 0
        ingresos_brutos= inc.get("ventas_brutas", ingresos) or ingresos
        devoluciones   = inc.get("devoluciones", 0) or 0

        activo_total   = bal.get("activo_total", 0) or 0
        pasivo_total   = bal.get("pasivo_total", 0) or 0
        capital        = bal.get("capital_contable", 0) or 0
        activo_cp      = bal.get("activo_circulante", 0) or 0
        pasivo_cp      = bal.get("pasivo_circulante", 0) or 0
        inventario     = bal.get("inventario", 0) or 0
        bancos         = bal.get("efectivo", bal.get("bancos", 0)) or 0

        margen_bruto   = util_bruta / ingresos if ingresos else 0
        margen_ebitda  = ebitda / ingresos if ingresos else 0
        margen_neto    = util_neta / ingresos if ingresos else 0
        liq_corriente  = activo_cp / pasivo_cp if pasivo_cp else 0
        prueba_acida   = (activo_cp - inventario) / pasivo_cp if pasivo_cp else 0
        razon_deuda    = pasivo_total / activo_total if activo_total else 0

        # ══════════════════════════════════════════════════════
        # HOJA 1: RESUMEN EJECUTIVO
        # ══════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Resumen Ejecutivo"
        ws1.sheet_view.showGridLines = False
        for col, w in [("A",3),("B",36),("C",16),("D",16),("E",16),("F",16),("G",16)]:
            ws1.column_dimensions[col].width = w

        ws1.row_dimensions[1].height = 8
        ws1.row_dimensions[2].height = 34
        hdr(ws1, 2, 2, f"{empresa.upper()} — REPORTE FINANCIERO EJECUTIVO", DARK_BLUE, WHITE, True, 14, span=6)
        ws1.row_dimensions[3].height = 18
        hdr(ws1, 3, 2, f"Período: {periodo}  |  RFC: {rfc}  |  Generado: {hoy}  |  TaxnFin · Claude Sonnet", MID_BLUE, WHITE, False, 10, span=6)
        ws1.row_dimensions[4].height = 8

        # KPI cards
        ws1.row_dimensions[5].height = 20
        hdr(ws1, 5, 2, "CIFRAS CLAVE DEL PERÍODO", DARK_BLUE, WHITE, True, 11, span=6)
        ws1.row_dimensions[6].height = 20
        kpi_hdrs = ["INGRESOS NETOS","UTIL. BRUTA","EBITDA","UTIL. NETA","MARGEN BRUTO","MARGEN NETO"]
        for i, t in enumerate(kpi_hdrs):
            hdr(ws1, 6, 2+i, t, GRAY_H, "404040", True, 9)
        ws1.row_dimensions[7].height = 28
        kpi_vals = [ingresos, util_bruta, ebitda, util_neta, margen_bruto, margen_neto]
        kpi_fmts = ['$#,##0','$#,##0','$#,##0','$#,##0','0.0%','0.0%']
        kpi_clrs = [MID_BLUE, GREEN, RED if ebitda<0 else GREEN, RED if util_neta<0 else GREEN, GREEN if margen_bruto>0.3 else ORANGE, RED if margen_neto<0 else GREEN]
        for i, (v, fmt, clr) in enumerate(zip(kpi_vals, kpi_fmts, kpi_clrs)):
            cell = ws1.cell(row=7, column=2+i, value=v)
            cell.font = Font(bold=True, color=WHITE, size=14, name="Arial")
            cell.fill = PatternFill("solid", fgColor=clr)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.number_format = fmt
        ws1.row_dimensions[8].height = 8

        # P&L table
        ws1.row_dimensions[9].height = 18
        hdr(ws1, 9, 2, "ESTADO DE RESULTADOS", DARK_BLUE, WHITE, True, 11, span=6)
        ws1.row_dimensions[10].height = 18
        hdr(ws1, 10, 2, "Concepto", MID_BLUE, WHITE, True, 10, "left")
        hdr(ws1, 10, 3, "Importe", MID_BLUE, WHITE, True, 10)
        hdr(ws1, 10, 4, "% Margen", MID_BLUE, WHITE, True, 10)
        hdr(ws1, 10, 5, "Diagnóstico", MID_BLUE, WHITE, True, 10, "left", span=3)

        pl_rows = [
            ("Ingresos brutos",         ingresos_brutos, None,          False, None,    None),
            ("(–) Devoluciones",        devoluciones,    None,          False, None,    LIGHT_RED if devoluciones else None),
            ("Ingresos netos",          ingresos,        None,          True,  MID_BLUE,None),
            ("(–) Costo de venta",      costo_venta,     None,          False, None,    None),
            ("Utilidad bruta",          util_bruta,      margen_bruto,  True,  GREEN,   None),
            ("(–) Gastos operativos",   gastos_op,       None,          False, None,    None),
            ("EBITDA",                  ebitda,          margen_ebitda, True,  RED if ebitda<0 else GREEN, None),
            ("(–) Gastos financieros",  gastos_fin,      None,          False, None,    None),
            ("Utilidad neta",           util_neta,       margen_neto,   True,  RED if util_neta<0 else GREEN, None),
        ]

        r = 11
        for concepto, importe, margen, bold, clr, bg in pl_rows:
            ws1.row_dimensions[r].height = 18
            fg = WHITE if clr in [MID_BLUE, GREEN, RED] else "000000"
            cell = ws1.cell(row=r, column=2, value=concepto)
            cell.font = Font(bold=bold, color=fg, size=10, name="Arial")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if clr: cell.fill = PatternFill("solid", fgColor=clr)
            elif bg: cell.fill = PatternFill("solid", fgColor=bg)

            cell2 = ws1.cell(row=r, column=3, value=importe)
            cell2.number_format = '$#,##0'
            cell2.font = Font(bold=bold, color=fg, size=10, name="Arial")
            cell2.alignment = Alignment(horizontal="right", vertical="center")
            if clr: cell2.fill = PatternFill("solid", fgColor=clr)
            elif bg: cell2.fill = PatternFill("solid", fgColor=bg)

            cell3 = ws1.cell(row=r, column=4, value=margen)
            if margen is not None:
                cell3.number_format = '0.0%'
                cell3.font = Font(bold=bold, color=fg, size=10, name="Arial")
                cell3.alignment = Alignment(horizontal="right", vertical="center")
                if clr: cell3.fill = PatternFill("solid", fgColor=clr)
            r += 1

        ws1.row_dimensions[r].height = 8; r += 1

        # Indicadores
        hdr(ws1, r, 2, "INDICADORES DE SALUD FINANCIERA", DARK_BLUE, WHITE, True, 11, span=6); ws1.row_dimensions[r].height = 18; r += 1
        hdr(ws1, r, 2, "Indicador", MID_BLUE, WHITE, True, 10, "left")
        hdr(ws1, r, 3, "Resultado", MID_BLUE, WHITE, True, 10)
        hdr(ws1, r, 4, "Nivel Saludable", MID_BLUE, WHITE, True, 10)
        hdr(ws1, r, 5, "Estado", MID_BLUE, WHITE, True, 10, "left", span=3)
        ws1.row_dimensions[r].height = 18; r += 1

        def estado(v, bueno, malo, fmt='0.00x', inv=False):
            if inv:
                clr = GREEN if v < malo else (RED if v > bueno else ORANGE)
            else:
                clr = GREEN if v > bueno else (RED if v < malo else ORANGE)
            return clr

        indicadores = [
            ("Liquidez corriente",    liq_corriente,  "0.00x", estado(liq_corriente, 1.5, 1.0), ">1.5x"),
            ("Prueba ácida",          prueba_acida,   "0.00x", estado(prueba_acida, 1.0, 0.5),  ">1.0x"),
            ("Razón de deuda",        razon_deuda,    "0.0%",  estado(razon_deuda, 0.6, 1.0, inv=True), "<60%"),
            ("Margen bruto",          margen_bruto,   "0.0%",  estado(margen_bruto, 0.30, 0.15), "30-65%"),
            ("Margen operativo",      margen_ebitda,  "0.0%",  estado(margen_ebitda, 0.05, 0.0), ">5%"),
            ("Margen neto",           margen_neto,    "0.0%",  estado(margen_neto, 0.03, 0.0),   ">3%"),
        ]

        for ind, resultado, fmt2, clr, nivel in indicadores:
            ws1.row_dimensions[r].height = 18
            if r % 2 == 0:
                for c in range(2, 8): ws1.cell(r, c).fill = PatternFill("solid", fgColor=GRAY_H)
            lbl(ws1, r, 2, ind)
            cell = ws1.cell(row=r, column=3, value=resultado)
            cell.number_format = fmt2
            cell.font = Font(bold=True, color=clr, size=11, name="Arial")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            lbl(ws1, r, 4, nivel, fg="595959")
            ws1.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
            cell2 = ws1.cell(row=r, column=5, value="●")
            cell2.font = Font(bold=True, color=clr, size=16, name="Arial")
            cell2.alignment = Alignment(horizontal="center", vertical="center")
            r += 1

        # AI Analysis si existe
        ai_summary = data.ai_analysis.get("summary", "") or data.ai_analysis.get("resumen", "")
        if ai_summary:
            ws1.row_dimensions[r].height = 8; r += 1
            hdr(ws1, r, 2, "ANÁLISIS EJECUTIVO IA", DARK_BLUE, WHITE, True, 11, span=6); ws1.row_dimensions[r].height = 18; r += 1
            ws1.row_dimensions[r].height = 80
            cell = ws1.cell(row=r, column=2, value=ai_summary)
            cell.font = Font(size=10, name="Arial", color="202020", italic=True)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
            r += 1

        ws1.freeze_panes = "C11"

        # ══════════════════════════════════════════════════════
        # HOJA 2: P&L DETALLADO
        # ══════════════════════════════════════════════════════
        ws2 = wb.create_sheet("P&L Detallado")
        ws2.sheet_view.showGridLines = False
        for col, w in [("A",3),("B",40),("C",18),("D",14)]:
            ws2.column_dimensions[col].width = w

        ws2.row_dimensions[1].height = 8
        ws2.row_dimensions[2].height = 30
        hdr(ws2, 2, 2, f"{empresa.upper()} — ESTADO DE RESULTADOS — {periodo}", DARK_BLUE, WHITE, True, 13, span=3)
        ws2.row_dimensions[3].height = 8
        ws2.row_dimensions[4].height = 18
        hdr(ws2, 4, 2, "Concepto", MID_BLUE, WHITE, True, 10, "left")
        hdr(ws2, 4, 3, "Importe MXN", MID_BLUE, WHITE, True, 10)
        hdr(ws2, 4, 4, "% Ingresos", MID_BLUE, WHITE, True, 10)

        pl_detail = [
            ("INGRESOS",               None,            None,           True,  MID_BLUE),
            ("Ingresos brutos",        ingresos_brutos, None,           False, None),
            ("(–) Devoluciones y NC",  devoluciones,    None,           False, LIGHT_RED if devoluciones else None),
            ("INGRESOS NETOS",         ingresos,        1.0,            True,  MID_BLUE),
            ("COSTO DE VENTA",         None,            None,           True,  MID_BLUE),
            ("Costo de venta",         costo_venta,     costo_venta/ingresos if ingresos else 0, False, None),
            ("UTILIDAD BRUTA",         util_bruta,      margen_bruto,   True,  GREEN),
            ("GASTOS OPERATIVOS",      None,            None,           True,  MID_BLUE),
            ("Gastos de operación",    gastos_op,       gastos_op/ingresos if ingresos else 0, False, None),
            ("UTILIDAD OPERATIVA (EBIT)", ebitda,       margen_ebitda,  True,  RED if ebitda<0 else GREEN),
            ("GASTOS FINANCIEROS",     None,            None,           True,  MID_BLUE),
            ("Intereses y comisiones", gastos_fin,      gastos_fin/ingresos if ingresos else 0, False, None),
            ("RESULTADO NETO",         util_neta,       margen_neto,    True,  RED if util_neta<0 else GREEN),
        ]

        r = 5
        for concepto, importe, margen, bold, clr in pl_detail:
            ws2.row_dimensions[r].height = 18
            fg = WHITE if clr in [MID_BLUE, GREEN, RED] else "000000"
            cell = ws2.cell(row=r, column=2, value=concepto)
            cell.font = Font(bold=bold, color=fg, size=10, name="Arial")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if clr: cell.fill = PatternFill("solid", fgColor=clr)

            if importe is not None:
                cell2 = ws2.cell(row=r, column=3, value=importe)
                cell2.number_format = '$#,##0'
                cell2.font = Font(bold=bold, color=fg, size=10, name="Arial")
                cell2.alignment = Alignment(horizontal="right", vertical="center")
                if clr: cell2.fill = PatternFill("solid", fgColor=clr)

            if margen is not None:
                cell3 = ws2.cell(row=r, column=4, value=margen)
                cell3.number_format = '0.0%'
                cell3.font = Font(bold=bold, color=fg, size=10, name="Arial")
                cell3.alignment = Alignment(horizontal="right", vertical="center")
                if clr: cell3.fill = PatternFill("solid", fgColor=clr)
            r += 1

        ws2.freeze_panes = "C5"

        # ══════════════════════════════════════════════════════
        # HOJA 3: BALANCE GENERAL
        # ══════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Balance General")
        ws3.sheet_view.showGridLines = False
        for col, w in [("A",3),("B",40),("C",18),("D",14)]:
            ws3.column_dimensions[col].width = w

        ws3.row_dimensions[1].height = 8
        ws3.row_dimensions[2].height = 30
        hdr(ws3, 2, 2, f"{empresa.upper()} — BALANCE GENERAL — {periodo}", DARK_BLUE, WHITE, True, 13, span=3)
        ws3.row_dimensions[3].height = 8
        ws3.row_dimensions[4].height = 18
        hdr(ws3, 4, 2, "Concepto", MID_BLUE, WHITE, True, 10, "left")
        hdr(ws3, 4, 3, "Importe MXN", MID_BLUE, WHITE, True, 10)
        hdr(ws3, 4, 4, "% Activo Total", MID_BLUE, WHITE, True, 10)

        activo_fijo = bal.get("activo_fijo", activo_total - activo_cp) or 0
        pasivo_lp   = bal.get("pasivo_largo_plazo", pasivo_total - pasivo_cp) or 0

        bal_rows = [
            ("ACTIVO",                None,        None, True,  MID_BLUE),
            ("Activo circulante",     activo_cp,   activo_cp/activo_total if activo_total else 0, False, LIGHT_BLUE),
            ("  Efectivo y bancos",   bancos,      bancos/activo_total if activo_total else 0, False, None),
            ("  Inventario",          inventario,  inventario/activo_total if activo_total else 0, False, None),
            ("Activo fijo",           activo_fijo, activo_fijo/activo_total if activo_total else 0, False, None),
            ("TOTAL ACTIVO",          activo_total, 1.0, True, DARK_BLUE),
            ("PASIVO",                None,        None, True,  MID_BLUE),
            ("Pasivo circulante",     pasivo_cp,   pasivo_cp/activo_total if activo_total else 0, False, LIGHT_RED),
            ("Pasivo largo plazo",    pasivo_lp,   pasivo_lp/activo_total if activo_total else 0, False, None),
            ("TOTAL PASIVO",          pasivo_total, pasivo_total/activo_total if activo_total else 0, True, RED),
            ("CAPITAL CONTABLE",      capital,      capital/activo_total if activo_total else 0, True, GREEN if capital>0 else RED),
            ("PASIVO + CAPITAL",      pasivo_total+capital, 1.0, True, DARK_BLUE),
        ]

        r = 5
        for concepto, importe, pct2, bold, clr in bal_rows:
            ws3.row_dimensions[r].height = 18
            fg = WHITE if clr in [MID_BLUE, GREEN, RED, DARK_BLUE] else "000000"
            cell = ws3.cell(row=r, column=2, value=concepto)
            cell.font = Font(bold=bold, color=fg, size=10, name="Arial")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if clr: cell.fill = PatternFill("solid", fgColor=clr)
            elif r % 2 == 0: cell.fill = PatternFill("solid", fgColor=GRAY_H)

            if importe is not None:
                cell2 = ws3.cell(row=r, column=3, value=importe)
                cell2.number_format = '$#,##0'
                cell2.font = Font(bold=bold, color=fg, size=10, name="Arial")
                cell2.alignment = Alignment(horizontal="right", vertical="center")
                if clr: cell2.fill = PatternFill("solid", fgColor=clr)
                elif r % 2 == 0: cell2.fill = PatternFill("solid", fgColor=GRAY_H)

            if pct2 is not None:
                cell3 = ws3.cell(row=r, column=4, value=pct2)
                cell3.number_format = '0.0%'
                cell3.font = Font(bold=bold, color=fg, size=10, name="Arial")
                cell3.alignment = Alignment(horizontal="right", vertical="center")
                if clr: cell3.fill = PatternFill("solid", fgColor=clr)
                elif r % 2 == 0: cell3.fill = PatternFill("solid", fgColor=GRAY_H)
            r += 1

        ws3.freeze_panes = "C5"

        # ══════════════════════════════════════════════════════
        # HOJA 4: INDICADORES FINANCIEROS
        # ══════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Indicadores Financieros")
        ws4.sheet_view.showGridLines = False
        for col, w in [("A",3),("B",34),("C",16),("D",16),("E",3),("F",55)]:
            ws4.column_dimensions[col].width = w

        ws4.row_dimensions[1].height = 8
        ws4.row_dimensions[2].height = 30
        hdr(ws4, 2, 2, f"{empresa.upper()} — ANÁLISIS DE INDICADORES — {periodo}", DARK_BLUE, WHITE, True, 13, span=5)
        ws4.row_dimensions[3].height = 8

        secciones_ind = [
            ("LIQUIDEZ", [
                ("Liquidez corriente",  liq_corriente,         "0.00x", ">1.5x",  GREEN if liq_corriente>1.5 else RED,  "Por cada $1 de deuda CP la empresa tiene activos para cubrirla."),
                ("Prueba ácida",        prueba_acida,          "0.00x", ">1.0x",  GREEN if prueba_acida>1.0 else RED,   "Capacidad de pago sin depender del inventario."),
                ("Razón de efectivo",   bancos/pasivo_cp if pasivo_cp else 0, "0.00x", ">0.2x", GREEN if (bancos/pasivo_cp if pasivo_cp else 0)>0.2 else RED, "Efectivo disponible vs deudas inmediatas."),
                ("Capital de trabajo",  activo_cp-pasivo_cp,   "$#,##0", ">$0",   GREEN if activo_cp>pasivo_cp else RED, "Excedente de recursos de corto plazo para operar."),
            ]),
            ("ENDEUDAMIENTO", [
                ("Razón de deuda",      razon_deuda,           "0.0%",  "<60%",   GREEN if razon_deuda<0.6 else RED,    "Proporción del activo financiada por terceros."),
                ("Deuda / Capital",     pasivo_total/capital if capital and capital>0 else 0, "0.00x", "<2.0x", GREEN if 0<(pasivo_total/capital if capital else 0)<2 else RED, "Apalancamiento: cuánta deuda por cada peso de capital."),
            ]),
            ("RENTABILIDAD", [
                ("Margen bruto",        margen_bruto,          "0.0%",  "30-65%", GREEN if margen_bruto>0.3 else (ORANGE if margen_bruto>0.15 else RED), "Porcentaje de ingresos que queda tras el costo del producto."),
                ("Margen operativo",    margen_ebitda,         "0.0%",  ">5%",    GREEN if margen_ebitda>0.05 else RED, "Porcentaje tras cubrir costos y gastos operativos."),
                ("Margen neto",         margen_neto,           "0.0%",  ">3%",    GREEN if margen_neto>0.03 else RED,   "Margen final tras todos los gastos e impuestos."),
                ("ROA",                 (util_neta*12)/activo_total if activo_total else 0, "0.0%", ">5%", GREEN if ((util_neta*12)/activo_total if activo_total else 0)>0.05 else RED, "Rendimiento anualizado sobre activos totales."),
            ]),
        ]

        r = 4
        for seccion, inds in secciones_ind:
            ws4.row_dimensions[r].height = 20
            hdr(ws4, r, 2, seccion, DARK_BLUE, WHITE, True, 11, span=5); r += 1
            ws4.row_dimensions[r].height = 18
            for c, t in [(2,"Indicador"),(3,"Resultado"),(4,"Nivel Saludable"),(5,""),(6,"Interpretación")]:
                hdr(ws4, r, c, t, MID_BLUE, WHITE, True, 10, "left" if c in [2,6] else "center"); r += 1
            for ind, val2, fmt2, nivel, clr, interp in inds:
                ws4.row_dimensions[r].height = 28
                lbl(ws4, r, 2, ind, bold=True)
                cell = ws4.cell(row=r, column=3, value=val2)
                cell.number_format = fmt2
                cell.font = Font(bold=True, size=12, color=clr, name="Arial")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                lbl(ws4, r, 4, nivel, fg="595959")
                ws4.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
                cell2 = ws4.cell(row=r, column=5, value="●")
                cell2.font = Font(bold=True, size=16, color=clr, name="Arial")
                cell2.alignment = Alignment(horizontal="center", vertical="center")
                lbl(ws4, r, 6, interp, italic=True, fg="404040")
                ws4.cell(r, 6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                if r % 2 == 0:
                    for c in range(2, 7):
                        if not ws4.cell(r,c).fill or ws4.cell(r,c).fill.fgColor.rgb == "00000000":
                            ws4.cell(r,c).fill = PatternFill("solid", fgColor=GRAY_H)
                r += 1
            ws4.row_dimensions[r].height = 6; r += 1

        # ══════════════════════════════════════════════════════
        # HOJA 5: TENDENCIAS
        # ══════════════════════════════════════════════════════
        ws5 = wb.create_sheet("Tendencias")
        ws5.sheet_view.showGridLines = False
        for col, w in [("A",3),("B",16),("C",16),("D",16),("E",16),("F",16),("G",16)]:
            ws5.column_dimensions[col].width = w

        ws5.row_dimensions[1].height = 8
        ws5.row_dimensions[2].height = 30
        hdr(ws5, 2, 2, f"{empresa.upper()} — TENDENCIAS HISTÓRICAS — {periodo}", DARK_BLUE, WHITE, True, 13, span=6)
        ws5.row_dimensions[3].height = 8
        ws5.row_dimensions[4].height = 18

        if data.trends:
            cols_t = ["Período","Ingresos","Util. Bruta","EBITDA","Util. Neta","Mg. Bruto"]
            for i, t in enumerate(cols_t):
                hdr(ws5, 4, 2+i, t, DARK_BLUE, WHITE, True, 10, "left" if i==0 else "center")
            r = 5
            for tr in data.trends:
                ws5.row_dimensions[r].height = 18
                bg = GRAY_H if r % 2 == 0 else WHITE
                periodo_t = tr.get("periodo", tr.get("period", ""))
                ing_t     = tr.get("ingresos", tr.get("ingresos_netos", 0)) or 0
                ub_t      = tr.get("utilidad_bruta", 0) or 0
                eb_t      = tr.get("ebitda", tr.get("utilidad_operativa", 0)) or 0
                un_t      = tr.get("utilidad_neta", tr.get("resultado_neto", 0)) or 0
                mb_t      = ub_t/ing_t if ing_t else 0

                for c, v2, fmt2 in [(2,periodo_t,None),(3,ing_t,'$#,##0'),(4,ub_t,'$#,##0'),(5,eb_t,'$#,##0'),(6,un_t,'$#,##0'),(7,mb_t,'0.0%')]:
                    cell = ws5.cell(row=r, column=c, value=v2)
                    if fmt2: cell.number_format = fmt2
                    cell.font = Font(size=10, name="Arial", color=RED if isinstance(v2, (int,float)) and v2 < 0 else "000000")
                    cell.alignment = Alignment(horizontal="right" if c > 2 else "left", vertical="center")
                    if bg != WHITE: cell.fill = PatternFill("solid", fgColor=bg)
                r += 1
        else:
            ws5.row_dimensions[4].height = 30
            lbl(ws5, 4, 2, "No hay datos históricos disponibles para el período seleccionado.", italic=True, fg="808080")

        ws5.freeze_panes = "C5"

        # ── Guardar en buffer ──────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        empresa_safe = data.empresa.replace(" ", "_").replace("/", "-")
        filename = f"Reporte_Financiero_{empresa_safe}_{data.periodo}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        logger.error(f"Error generando Excel corporativo: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generando Excel: {str(e)}")
